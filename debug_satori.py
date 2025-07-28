#!/usr/bin/env python3
"""
SATORI実績ファイルのシート分析スクリプト
"""

import pandas as pd
import openpyxl

def analyze_satori_file():
    file_path = r'C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ISP支払通知書\2024\202408\【株式会社アウトワード御中】SATORI実績_2408月次レポート.xlsx'

    # Excelファイルを読み込み
    wb = openpyxl.load_workbook(file_path, data_only=True)

    print('=== 従量実績シート - 全C列データ（空白以外） ===')
    try:
        juryou_sheet = wb['従量実績']
        juryou_c_values = set()
        for row in range(1, 100):  # 最初の100行を調査
            cell_value = juryou_sheet.cell(row=row, column=3).value
            if cell_value is not None and str(cell_value).strip() != '':
                juryou_c_values.add(str(cell_value))
        
        print('従量実績シートのC列ユニーク値:')
        for value in sorted(juryou_c_values):
            print(f'  - {value}')
        print(f'総数: {len(juryou_c_values)}')
                
    except Exception as e:
        print(f'従量実績シートの読み込みエラー: {e}')

    print('\n=== docomo占いシート - 全C列データ（空白以外） ===')
    try:
        docomo_sheet = wb['docomo占い']
        docomo_c_values = set()
        for row in range(1, 100):  # 最初の100行を調査
            cell_value = docomo_sheet.cell(row=row, column=3).value
            if cell_value is not None and str(cell_value).strip() != '':
                docomo_c_values.add(str(cell_value))
        
        print('docomo占いシートのC列ユニーク値:')
        for value in sorted(docomo_c_values):
            print(f'  - {value}')
        print(f'総数: {len(docomo_c_values)}')
        
    except Exception as e:
        print(f'docomo占いシートの読み込みエラー: {e}')

    print('\n=== 一致分析 ===')
    # 一致するもの
    matching = juryou_c_values & docomo_c_values
    print(f'両方のシートに存在するコンテンツ（{len(matching)}件）:')
    for value in sorted(matching):
        print(f'  ✓ {value}')

    # 従量実績のみ
    juryou_only = juryou_c_values - docomo_c_values
    print(f'\n従量実績シートのみに存在するコンテンツ（{len(juryou_only)}件）:')
    for value in sorted(juryou_only):
        print(f'  × {value}')

    # docomo占いのみ  
    docomo_only = docomo_c_values - juryou_c_values
    print(f'\ndocomo占いシートのみに存在するコンテンツ（{len(docomo_only)}件）:')
    for value in sorted(docomo_only):
        print(f'  × {value}')

    wb.close()

if __name__ == "__main__":
    analyze_satori_file()