#!/usr/bin/env python3
"""
AC列の内容を確認するスクリプト
"""

import openpyxl
from pathlib import Path

def check_ac_column():
    """aokiテンプレートと生成ファイルのAC列を確認"""
    template_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\コンテンツ関連支払明細書フォーマット\aoki.xlsx"
    generated_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    print("=== AC列の確認 ===")
    
    # テンプレートファイルのAC列を確認
    if Path(template_path).exists():
        print("\n--- テンプレートファイル aoki.xlsx ---")
        wb_template = openpyxl.load_workbook(template_path, data_only=False)
        ws_template = wb_template.active
        
        for row in range(20, 30):
            cell_ref = f"AC{row}"
            cell = ws_template[cell_ref]
            if cell.value is not None and cell.value != "":
                print(f"  {cell_ref}: {cell.value}")
    
    # 生成ファイルのAC列を確認
    if Path(generated_path).exists():
        print("\n--- 生成ファイル 202505_aoki.xlsx ---")
        wb_generated = openpyxl.load_workbook(generated_path, data_only=False)
        ws_generated = wb_generated.active
        
        for row in range(20, 30):
            cell_ref = f"AC{row}"
            cell = ws_generated[cell_ref]
            if cell.value is not None and cell.value != "":
                print(f"  {cell_ref}: {cell.value}")
            
            # データ版でも確認
            wb_data = openpyxl.load_workbook(generated_path, data_only=True)
            ws_data = wb_data.active
            data_cell = ws_data[cell_ref]
            if data_cell.value is not None and data_cell.value != "" and data_cell.value != 0:
                print(f"  {cell_ref} データ値: {data_cell.value}")
    
    # AE列の計算を手動でやってみる
    print("\n--- 手動計算 ---")
    if Path(generated_path).exists():
        wb = openpyxl.load_workbook(generated_path, data_only=True)
        ws = wb.active
        
        # aokiの料率は8%
        aoki_rate = 8.0
        
        total = 0
        for row in [23, 25, 26]:  # データがある行
            y_val = ws[f'Y{row}'].value
            if y_val and isinstance(y_val, (int, float)):
                calculated = y_val * aoki_rate * 0.01
                print(f"  行{row}: Y{row}={y_val} × {aoki_rate}% = {calculated}")
                total += calculated
        
        print(f"合計: {total}")
        print(f"期待値（58+81+163）: {58+81+163}")
        print(f"差分: {total - 302}")

if __name__ == '__main__':
    check_ac_column()