#\!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AE23-AE58セルの数式を詳細に分析
"""

import openpyxl
import glob
import os

def analyze_ae_formulas(file_path):
    """AE23-AE58セルの数式を分析"""
    print(f"=== {os.path.basename(file_path)} のAE23-AE58分析 ===")
    
    try:
        # 数式ブックと計算値ブックの両方を読み込み
        workbook_formula = openpyxl.load_workbook(file_path, data_only=False)
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        
        worksheet_formula = workbook_formula.active
        worksheet_data = workbook_data.active
        
        print(f"AE23-AE58の数式と値:")
        
        for row in range(23, 59):  # 23-58行目
            ae_formula = worksheet_formula.cell(row=row, column=31)  # AE列 = 31列目
            ae_data = worksheet_data.cell(row=row, column=31)
            
            if ae_formula.value is not None:
                print(f"  AE{row}:")
                print(f"    数式: {ae_formula.value}")
                print(f"    計算値: {ae_data.value}")
                print(f"    型: {type(ae_data.value)}")
                
                # 数式の場合、さらに参照先を確認
                if isinstance(ae_formula.value, str) and ae_formula.value.startswith('='):
                    print(f"    -> これは数式です")
                print()
        
        workbook_formula.close()
        workbook_data.close()
        
    except Exception as e:
        print(f"エラー: {e}")

def main():
    # テスト用に1つのファイルを分析
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    
    # 202508フォルダから1つのファイルを選択
    test_folder = os.path.join(royalty_dir, "202508")
    if os.path.exists(test_folder):
        xlsx_files = glob.glob(os.path.join(test_folder, "*.xlsx"))
        if xlsx_files:
            # 最初のファイルをテスト
            analyze_ae_formulas(xlsx_files[0])
        else:
            print("No xlsx files found")
    else:
        print("Test folder not found")

if __name__ == "__main__":
    main()
