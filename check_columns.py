#\!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Y列とAC列の値を確認
"""

import openpyxl
import glob
import os

def check_y_and_ac_columns(file_path):
    """Y列とAC列の値を確認"""
    print(f"=== {os.path.basename(file_path)} のY列とAC列分析 ===")
    
    try:
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        worksheet_data = workbook_data.active
        
        print(f"Y列とAC列の値（23-58行目）:")
        
        for row in range(23, 59):  # 23-58行目
            y_value = worksheet_data.cell(row=row, column=25).value  # Y列 = 25列目
            ac_value = worksheet_data.cell(row=row, column=29).value  # AC列 = 29列目
            
            if (y_value is not None and y_value != "" and y_value != 0) or (ac_value is not None and ac_value != "" and ac_value != 0):
                calculated_ae = ""
                if isinstance(y_value, (int, float)) and isinstance(ac_value, (int, float)) and y_value != 0 and ac_value != 0:
                    calculated_ae = f" -> AE{row}計算値: {y_value * ac_value * 0.01}"
                
                print(f"  行{row}: Y{row}={y_value}, AC{row}={ac_value}{calculated_ae}")
        
        workbook_data.close()
        
    except Exception as e:
        print(f"エラー: {e}")

def main():
    # テスト用に1つのファイルを分析
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    
    # 202508フォルダから1つのファイルを選択
    test_folder = os.path.join(royalty_dir, "202508")
    if os.path.exists(test_folder):
        xlsx_files = glob.glob(os.path.join(test_folder, "*.xlsx"))
        if xlsx_files:
            # 最初のファイルをテスト
            check_y_and_ac_columns(xlsx_files[0])
        else:
            print("No xlsx files found")
    else:
        print("Test folder not found")

if __name__ == "__main__":
    main()
