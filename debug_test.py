import openpyxl
import os

file_path = r'C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202507\202507_aiga.xlsx'
print(f'Testing file: {file_path}')
print(f'File exists: {os.path.exists(file_path)}')

if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    worksheet = workbook.active
    print(f'Worksheet: {worksheet.title}')
    print(f'Max row: {worksheet.max_row}, Max column: {worksheet.max_column}')
    
    ae19_value = worksheet.cell(row=19, column=31).value
    print(f'AE19 value: {ae19_value}')
    
    for row in range(17, 22):
        for col in range(29, 35):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and isinstance(cell_value, (int, float)) and cell_value > 0:
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f'{col_letter}{row}: {cell_value}')
    
    workbook.close()

