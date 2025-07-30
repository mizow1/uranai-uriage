#!/usr/bin/env python3
"""
全テンプレートファイルのAE列数式を修正するスクリプト
"""

import openpyxl
from pathlib import Path
import shutil
import re
from typing import List

def get_all_template_files() -> List[Path]:
    """全テンプレートファイルのリストを取得"""
    template_dir = Path(r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\コンテンツ関連支払明細書フォーマット")
    
    # .xlsxファイルを全て取得
    excel_files = list(template_dir.glob("*.xlsx"))
    return sorted(excel_files)

def fix_template_file(template_path: Path) -> bool:
    """単一のテンプレートファイルを修正"""
    try:
        print(f"修正中: {template_path.name}")
        
        # バックアップディレクトリ
        backup_dir = template_path.parent / "bk"
        backup_dir.mkdir(exist_ok=True)
        
        # バックアップファイル名
        backup_name = template_path.stem + "-bk.xlsx"
        backup_path = backup_dir / backup_name
        
        # バックアップを作成（既に存在する場合はスキップ）
        if not backup_path.exists():
            shutil.copy2(template_path, backup_path)
            print(f"  バックアップ作成: {backup_name}")
        
        # テンプレートファイルを開く
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        modified_count = 0
        
        # 修正対象の行範囲（23-58行）
        for row in range(23, 59):
            cell_ref = f"AE{row}"
            cell = ws[cell_ref]
            
            if cell.value and str(cell.value).startswith('='):
                old_formula = str(cell.value)
                
                # ROUND関数がすでに含まれている場合はスキップ
                if "ROUND(" in old_formula:
                    continue
                
                # 数式を修正：ROUND関数で整数に丸める
                # パターン: Y{row}*AC{row}*0.01 → ROUND(Y{row}*AC{row}*0.01,0)
                pattern = rf'(Y{row}\*AC{row}\*0\.01)(\)\)?)$'
                
                def replace_func(match):
                    calculation = match.group(1)
                    ending = match.group(2) if match.group(2) else "))"
                    return f"ROUND({calculation},0){ending}"
                
                new_formula = re.sub(pattern, replace_func, old_formula)
                
                if new_formula != old_formula:
                    ws[cell_ref].value = new_formula
                    modified_count += 1
        
        if modified_count > 0:
            # ファイルを保存
            wb.save(template_path)
            print(f"  修正完了: {modified_count}個のセルを更新")
        else:
            print(f"  修正不要: すでにROUND関数が適用済み")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  エラー: {e}")
        return False

def verify_template_file(template_path: Path) -> bool:
    """テンプレートファイルの修正を検証"""
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # いくつかのセルでROUND関数が適用されているか確認
        for row in [23, 24, 25]:
            cell_ref = f"AE{row}"
            cell = ws[cell_ref]
            
            if cell.value and str(cell.value).startswith('='):
                formula = str(cell.value)
                if f"Y{row}*AC{row}*0.01" in formula and "ROUND(" not in formula:
                    wb.close()
                    return False
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"検証エラー: {e}")
        return False

def main():
    """メイン処理"""
    print("=== 全テンプレートファイルのAE列数式修正 ===")
    
    # 全テンプレートファイルを取得
    template_files = get_all_template_files()
    print(f"対象ファイル数: {len(template_files)}")
    
    success_count = 0
    error_count = 0
    
    # 各ファイルを修正
    for template_path in template_files:
        if fix_template_file(template_path):
            success_count += 1
        else:
            error_count += 1
    
    print(f"\n=== 修正結果 ===")
    print(f"成功: {success_count}ファイル")
    print(f"エラー: {error_count}ファイル")
    
    # 検証
    if success_count > 0:
        print(f"\n=== 修正内容の検証 ===")
        verify_success = 0
        verify_error = 0
        
        for template_path in template_files[:5]:  # 最初の5つのファイルで検証
            if verify_template_file(template_path):
                verify_success += 1
                print(f"✅ {template_path.name}: 修正済み")
            else:
                verify_error += 1
                print(f"❌ {template_path.name}: 未修正")
        
        print(f"\n検証結果: {verify_success}成功 / {verify_error}エラー")

if __name__ == '__main__':
    main()