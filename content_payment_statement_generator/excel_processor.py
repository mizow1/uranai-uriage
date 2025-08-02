"""
Excelファイル処理モジュール

Excelテンプレートの複製と明細データの書き込みを行います。
"""

import shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import logging
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .config_manager import ConfigManager
from .data_models import SalesRecord


class ExcelProcessor:
    """Excelファイル処理クラス"""
    
    def __init__(self, config_manager: ConfigManager):
        """Excel処理クラスを初期化"""
        self.config = config_manager
        self.logger = logging.getLogger(__name__)
    
    def copy_template(self, template_name: str, output_path: str, target_month: str, content_name: str = None) -> str:
        """テンプレートファイルを指定の出力パスに複製"""
        try:
            # テンプレートファイルのパスを取得
            template_path = self.config.get_template_file_by_name(template_name)
            
            if not template_path or not Path(template_path).exists():
                raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_name}")
            
            # 出力ファイル名を生成
            template_file = Path(template_path)
            if content_name:
                # コンテンツ名が指定されている場合: YYYYMM_content.xlsx
                output_filename = f"{target_month}_{content_name}.xlsx"
            else:
                # コンテンツ名が指定されていない場合: YYYYMM_元ファイル名
                output_filename = f"{target_month}_{template_file.name}"
            output_file_path = Path(output_path) / output_filename
            
            # ディレクトリを作成
            output_file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # ファイルを複製
            shutil.copy2(template_path, output_file_path)
            
            self.logger.info(f"テンプレートファイルを複製しました: {output_file_path}")
            return str(output_file_path)
            
        except Exception as e:
            self.logger.error(f"テンプレートファイル複製エラー: {e}")
            raise
    
    def write_payment_date(self, workbook_path: str, target_month: str, template_name: str) -> None:
        """S3セルに適切な日付を記入"""
        try:
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(workbook_path)
            worksheet = workbook.active
            
            # 対象年月を解析
            year = int(target_month[:4])
            month = int(target_month[4:])
            
            # テンプレート名から判定してS3セルに記入する日付を決定
            template_lower = template_name.lower()
            target_templates = ['epc', 'gaia', 'ichild', 'macalon', 'mermaid', 'shape', 'shintaku']
            
            if any(template in template_lower for template in target_templates):
                # 前月末日を計算
                if month == 1:
                    prev_year = year - 1
                    prev_month = 12
                else:
                    prev_year = year
                    prev_month = month - 1
                
                # 前月末日を取得
                if prev_month in [1, 3, 5, 7, 8, 10, 12]:
                    last_day = 31
                elif prev_month in [4, 6, 9, 11]:
                    last_day = 30
                else:  # 2月
                    # うるう年判定
                    if (prev_year % 4 == 0 and prev_year % 100 != 0) or (prev_year % 400 == 0):
                        last_day = 29
                    else:
                        last_day = 28
                
                date_to_write = datetime(prev_year, prev_month, last_day)
                self.logger.info(f"対象テンプレート {template_name}: S3セルに前月末日を記入: {date_to_write.strftime('%Y年%m月%d日')}")
            else:
                # その他のテンプレートは従来通り対象年月の5日
                date_to_write = datetime(year, month, 5)
                self.logger.info(f"通常テンプレート {template_name}: S3セルに5日を記入: {date_to_write.strftime('%Y年%m月%d日')}")
            
            # S3セルに日付を設定（yyyy年m月d日フォーマット）
            formatted_date = date_to_write.strftime('%Y年%m月%d日')
            worksheet['S3'] = formatted_date
            
            # ファイルを保存
            workbook.save(workbook_path)
            
            self.logger.info(f"日付を設定しました: S3セル = {formatted_date}")
            
        except Exception as e:
            self.logger.error(f"日付設定エラー: {e}")
            raise
        finally:
            if 'workbook' in locals():
                workbook.close()
    
    def write_statement_details(self, workbook_path: str, sales_records: List[SalesRecord], processing_month: str) -> None:
        """23行目以降に明細データを書き込み"""
        try:
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(workbook_path)
            worksheet = workbook.active
            
            # W21セルが「件数」かチェック
            w21_cell = worksheet['W21']
            has_count_column = (w21_cell.value and str(w21_cell.value).strip() == '件数')
            
            self.logger.info(f"W21セルの値: '{w21_cell.value}', 件数列有効: {has_count_column}")
            
            # 開始行（23行目）
            start_row = 23
            
            for i, record in enumerate(sales_records):
                row_num = start_row + i
                
                # 各列にデータを設定
                self._write_record_to_row(worksheet, row_num, record, processing_month, has_count_column)
            
            # ファイルを保存
            workbook.save(workbook_path)
            
            self.logger.info(f"明細データを書き込みました: {len(sales_records)}件")
            
        except Exception as e:
            self.logger.error(f"明細データ書き込みエラー: {e}")
            raise
        finally:
            if 'workbook' in locals():
                workbook.close()
    
    def _write_record_to_row(self, worksheet: Worksheet, row_num: int, record: SalesRecord, processing_month: str, has_count_column: bool = False) -> None:
        """1つのSalesRecordを指定行に書き込み"""
        try:
            # 要求仕様に基づく新しいルール:
            # A列：処理開始時に指定した年月
            # D列：プラットフォーム名  
            # G列：コンテンツ名
            # M列：target_month.csvのC列の数値分、対象年月からマイナスした年月
            # S列：該当年月の「実績」額
            # Y列：該当年月の「情報提供料」額
            # W列：件数（W21セルが「件数」の場合のみ、amebaとmedibaのみ対象）
            
            # セルへの書き込み（マージセルの場合は上位セルに書き込む）
            self._safe_write_cell(worksheet, f'A{row_num}', processing_month)       # A列：処理開始時に指定した年月
            self._safe_write_cell(worksheet, f'D{row_num}', record.platform)       # D列：プラットフォーム名
            self._safe_write_cell(worksheet, f'G{row_num}', record.content_name)   # G列：コンテンツ名
            self._safe_write_cell(worksheet, f'M{row_num}', record.target_month)   # M列：マイナスした年月
            self._safe_write_cell(worksheet, f'S{row_num}', record.performance)    # S列：実績額
            self._safe_write_cell(worksheet, f'Y{row_num}', record.information_fee) # Y列：情報提供料額
            
            # W列：件数（条件に合致する場合のみ）
            if has_count_column and record.platform.lower() in ['ameba', 'mediba'] and record.sales_count > 0:
                self._safe_write_cell(worksheet, f'W{row_num}', record.sales_count)
                self.logger.debug(f"W{row_num}に件数を記入: {record.sales_count} ({record.platform})")
            
            # AC列（料率）は数式が存在する場合は保持し、なければ値を設定
            self._write_rate_cell(worksheet, f'AC{row_num}', record.rate)
            
            self.logger.debug(f"行 {row_num} にレコードを書き込みました: {record.content_name}")
            
        except Exception as e:
            self.logger.error(f"レコード書き込みエラー (行 {row_num}): {e}")
            raise
    
    def _safe_write_cell(self, worksheet: Worksheet, cell_address: str, value) -> None:
        """セルに安全に値を書き込み（マージセル対応）"""
        try:
            # 直接的なアプローチ：マージセルエラーをキャッチして処理
            try:
                worksheet[cell_address].value = value
                self.logger.debug(f"セル {cell_address} に正常に書き込みました: {value}")
            except Exception as write_error:
                if "'MergedCell' object attribute 'value' is read-only" in str(write_error):
                    # マージセルの場合の処理
                    self._handle_merged_cell_write(worksheet, cell_address, value)
                else:
                    # 他のエラーの場合
                    self.logger.warning(f"セル {cell_address} への書き込みエラー: {write_error}")
            
        except Exception as e:
            self.logger.warning(f"セル {cell_address} への書き込み処理エラー: {e}")
    
    def _handle_merged_cell_write(self, worksheet: Worksheet, cell_address: str, value) -> None:
        """マージセル書き込み処理"""
        try:
            # マージされた範囲を検索
            from openpyxl.utils import coordinate_to_tuple
            row, col = coordinate_to_tuple(cell_address)
            
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col <= merged_range.max_col:
                    # マージ範囲の左上セルに書き込み
                    top_left_address = f"{merged_range.start_cell.column_letter}{merged_range.start_cell.row}"
                    worksheet[top_left_address].value = value
                    self.logger.debug(f"マージセル {cell_address} の代わりに {top_left_address} に書き込みました: {value}")
                    return
            
            # マージ範囲が見つからない場合は、強制的に書き込みを試行
            self.logger.warning(f"マージセル {cell_address} の範囲が特定できません。値: {value}")
            
        except Exception as e:
            self.logger.warning(f"マージセル処理エラー {cell_address}: {e}")
    
    def _write_rate_cell(self, worksheet: Worksheet, cell_address: str, rate_value: float) -> None:
        """料率セルの書き込み（数式がある場合は保持）"""
        try:
            cell = worksheet[cell_address]
            
            # セルに数式がある場合は保持
            if hasattr(cell, 'formula') and cell.formula:
                self.logger.debug(f"セル {cell_address} に数式が存在するため保持: {cell.formula}")
                return
            
            # セルの値が数式の場合も保持
            if isinstance(cell.value, str) and cell.value.startswith('='):
                self.logger.debug(f"セル {cell_address} に数式が存在するため保持: {cell.value}")
                return
            
            # 数式がない場合は値を設定
            self._safe_write_cell(worksheet, cell_address, rate_value)
            self.logger.debug(f"セル {cell_address} に料率値を設定: {rate_value}")
            
        except Exception as e:
            self.logger.warning(f"料率セル処理エラー {cell_address}: {e}")
            # エラーが発生した場合は通常の書き込みを試行
            self._safe_write_cell(worksheet, cell_address, rate_value)
    
    def calculate_payment_amount(self, performance: float, rate: float) -> float:
        """支払額を計算（実績 × 料率）"""
        try:
            return performance * rate
        except Exception as e:
            self.logger.error(f"支払額計算エラー: {e}")
            return 0.0
    
    def process_excel_file(
        self, 
        template_name: str, 
        sales_records: List[SalesRecord], 
        target_month: str,
        content_name: str = None
    ) -> str:
        """Excelファイルの完全処理（テンプレート複製 + データ書き込み）"""
        try:
            # 出力ディレクトリを取得
            year = target_month[:4]
            month = target_month[4:]
            output_dir = self.config.get_output_directory(year, month)
            
            # テンプレートを複製
            excel_path = self.copy_template(template_name, output_dir, target_month, content_name)
            
            # 支払日を設定
            self.write_payment_date(excel_path, target_month, template_name)
            
            # 明細データを書き込み（処理月を渡す）
            self.write_statement_details(excel_path, sales_records, target_month)
            
            self.logger.info(f"Excelファイル処理完了: {excel_path}")
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Excelファイル処理エラー: {e}")
            raise
    
    def validate_excel_structure(self, workbook_path: str) -> bool:
        """Excelファイルの構造を検証"""
        try:
            workbook = openpyxl.load_workbook(workbook_path)
            worksheet = workbook.active
            
            # 必要なセルの存在確認
            required_cells = ['S3']  # 支払日セル
            
            for cell in required_cells:
                if worksheet[cell].value is None:
                    self.logger.warning(f"必要なセル {cell} が空です")
                    return False
            
            self.logger.info("Excelファイル構造の検証が完了しました")
            return True
            
        except Exception as e:
            self.logger.error(f"Excelファイル構造検証エラー: {e}")
            return False
        finally:
            if 'workbook' in locals():
                workbook.close()