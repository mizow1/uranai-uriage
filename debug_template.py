#!/usr/bin/env python3
"""
元のテンプレートファイルと生成されたファイルを比較するスクリプト
"""

import pandas as pd
import openpyxl
from pathlib import Path

def compare_aoki_files():
    """aoki.xlsxテンプレートと202505_aoki.xlsxを比較"""
    template_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\コンテンツ関連支払明細書フォーマット\aoki.xlsx"
    generated_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    print("=== テンプレートファイルと生成ファイルの比較 ===")
    
    if not Path(template_path).exists():
        print(f"テンプレートファイルが見つかりません: {template_path}")
        return
    
    if not Path(generated_path).exists():
        print(f"生成ファイルが見つかりません: {generated_path}")
        return
    
    # テンプレートファイルを確認
    print("\n--- テンプレートファイル (aoki.xlsx) ---")
    wb_template = openpyxl.load_workbook(template_path, data_only=True)
    ws_template = wb_template.active
    
    # M19セルの確認（テンプレート）
    m19_template = ws_template['M19']
    print(f"テンプレートM19セル: {m19_template.value}")
    
    # 生成ファイルを確認
    print("\n--- 生成ファイル (202505_aoki.xlsx) ---")
    wb_generated = openpyxl.load_workbook(generated_path, data_only=True)
    ws_generated = wb_generated.active
    
    # M19セルの確認（生成ファイル）
    m19_generated = ws_generated['M19']
    print(f"生成ファイルM19セル: {m19_generated.value}")
    
    # 数式版も確認
    wb_template_formula = openpyxl.load_workbook(template_path, data_only=False)
    ws_template_formula = wb_template_formula.active
    wb_generated_formula = openpyxl.load_workbook(generated_path, data_only=False)
    ws_generated_formula = wb_generated_formula.active
    
    print(f"テンプレートM19セル数式: {ws_template_formula['M19'].value}")
    print(f"生成ファイルM19セル数式: {ws_generated_formula['M19'].value}")
    
    # Y列とAC列の比較（テンプレート）
    print(f"\n--- テンプレートファイルのY列とAC列（23-58行）---")
    template_data_found = False
    for row in range(23, 59):
        y_val = ws_template[f'Y{row}'].value
        ac_val = ws_template[f'AC{row}'].value
        
        if y_val is not None and ac_val is not None and y_val != "" and ac_val != "":
            print(f"  行{row}: Y{row}={y_val}, AC{row}={ac_val}")
            template_data_found = True
    
    if not template_data_found:
        print("  テンプレートにはY列とAC列にデータがありません")
    
    # Y列とAC列の比較（生成ファイル）
    print(f"\n--- 生成ファイルのY列とAC列（23-58行）---")
    generated_data_found = False
    for row in range(23, 59):
        y_val = ws_generated[f'Y{row}'].value
        ac_val = ws_generated[f'AC{row}'].value
        
        if y_val is not None and ac_val is not None and y_val != "" and ac_val != "":
            print(f"  行{row}: Y{row}={y_val}, AC{row}={ac_val}")
            generated_data_found = True
    
    if not generated_data_found:
        print("  生成ファイルにはY列とAC列にデータがありません")
    
    # 全体から58, 81, 163を探す
    print(f"\n--- 生成ファイルから58, 81, 163を検索 ---")
    target_values = [58, 81, 163]
    found_cells = []
    
    for row in range(1, ws_generated.max_row + 1):
        for col in range(1, ws_generated.max_column + 1):
            cell_value = ws_generated.cell(row=row, column=col).value
            if cell_value in target_values:
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                found_cells.append((cell_ref, cell_value))
                print(f"  {cell_ref}: {cell_value}")
    
    if found_cells:
        print(f"合計: {sum([v for _, v in found_cells])}")
        print(f"期待値: 302")
        print(f"差分: {sum([v for _, v in found_cells]) - 302}")
    else:
        print("58, 81, 163の値が見つかりませんでした")

if __name__ == '__main__':
    compare_aoki_files()