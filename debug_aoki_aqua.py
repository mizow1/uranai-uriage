#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aoki、aquaの売上計算をデバッグするスクリプト
"""

import os
import pandas as pd
import glob
from datetime import datetime
import openpyxl
import openpyxl.utils
from pathlib import Path
import re
import csv

def load_rate_data():
    """rate.csvを読み込み"""
    rate_csv_path = r"C:\Users\OW\pj\uriage\rate.csv"
    rate_data = pd.read_csv(rate_csv_path, encoding='utf-8')
    print(f"Rate data loaded: {len(rate_data)} records")
    return rate_data

def extract_name_from_filename(filename):
    """ファイル名から名称を抽出"""
    base_name = os.path.splitext(filename)[0]
    parts = base_name.split('_')
    yyyymm_pattern = re.compile(r'^\d{6}$')
    name_parts = [part for part in parts if not yyyymm_pattern.match(part)]
    
    if name_parts:
        return name_parts[0]
    return base_name

def read_sales_from_xlsx(file_path):
    """xlsxファイルから売上データを読み取り（デバッグ用）"""
    try:
        workbook_formula = openpyxl.load_workbook(file_path, data_only=False)
        worksheet_formula = workbook_formula.active
        
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        worksheet_data = workbook_data.active
        
        # AE19セル（31列目、19行目）の処理
        ae19_cell = worksheet_formula.cell(row=19, column=31)
        ae19_value = worksheet_data.cell(row=19, column=31).value
        
        print(f"  AE19セル - 数式: {ae19_cell.value}, 計算値: {ae19_value}")
        
        # M19の内容も確認
        m19_cell = worksheet_formula.cell(row=19, column=13)
        m19_value = worksheet_data.cell(row=19, column=13).value
        print(f"  M19セル - 数式: {m19_cell.value}, 計算値: {m19_value}")
        
        # S19、Y19の内容も確認
        s19_cell = worksheet_formula.cell(row=19, column=19)
        s19_value = worksheet_data.cell(row=19, column=19).value
        print(f"  S19セル - 数式: {s19_cell.value}, 計算値: {s19_value}")
        
        y19_cell = worksheet_formula.cell(row=19, column=25)
        y19_value = worksheet_data.cell(row=19, column=25).value
        print(f"  Y19セル - 数式: {y19_cell.value}, 計算値: {y19_value}")
        
        # AE19が数式の場合の詳細処理
        if ae19_cell.value and isinstance(ae19_cell.value, str) and ae19_cell.value.startswith('='):
            print(f"  AE19は数式です: {ae19_cell.value}")
            
            # AE23:AE58の範囲を詳細に調査
            ae_sum = 0
            found_values = []
            
            for row in range(23, 59):
                y_value = worksheet_data.cell(row=row, column=25).value  # Y列
                ac_value = worksheet_data.cell(row=row, column=29).value  # AC列
                ae_value = worksheet_data.cell(row=row, column=31).value  # AE列
                
                if isinstance(y_value, (int, float)) and y_value > 0:
                    # 実際のAC値を使用
                    if ac_value is None:
                        ac_value = 10 if y_value > 0 else 0
                    
                    # 実際のAE値を使用
                    if ae_value is not None:
                        ae_sum += ae_value
                    else:
                        calculated_ae = y_value * ac_value * 0.01
                        ae_sum += calculated_ae
                    
                    found_values.append(f"行{row}: Y={y_value}, AC={ac_value}, AE={ae_value}")
            
            if found_values:
                print(f"    売上データ詳細:")
                for value in found_values:
                    print(f"      {value}")
                print(f"    AE合計: {ae_sum}")
            
            # 最終計算値の確認
            if m19_value is not None:
                calculated_final = m19_value - (s19_value or 0) + (y19_value or 0)
                print(f"  最終計算値: M19({m19_value}) - S19({s19_value}) + Y19({y19_value}) = {calculated_final}")
                
                workbook_formula.close()
                workbook_data.close()
                return calculated_final
            elif ae19_value is not None:
                workbook_formula.close()
                workbook_data.close()
                return ae19_value
        
        elif ae19_value is not None and isinstance(ae19_value, (int, float)):
            print(f"  AE19は直接値: {ae19_value}")
            workbook_formula.close()
            workbook_data.close()
            return ae19_value
        
        workbook_formula.close()
        workbook_data.close()
        return 0
        
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return 0

def debug_specific_agents():
    """aoki、aquaの売上を詳細デバッグ"""
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    rate_data = load_rate_data()
    
    # 対象エージェント
    target_agents = ['aoki', 'aqua']
    
    # 対象月フォルダ
    target_months = ['202412', '202507', '202508']
    
    for month in target_months:
        print(f"\n=== {month}月の処理 ===")
        folder_path = os.path.join(royalty_dir, month)
        
        if not os.path.exists(folder_path):
            print(f"フォルダが存在しません: {folder_path}")
            continue
        
        xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        print(f"Excelファイル数: {len(xlsx_files)}")
        
        for file_path in xlsx_files:
            filename = os.path.basename(file_path)
            name = extract_name_from_filename(filename)
            
            # rate.csvで対応するエージェントを検索
            matching_row = rate_data[rate_data['名称'] == name]
            if not matching_row.empty:
                agent = matching_row.iloc[0]['エージェント']
                if pd.notna(agent) and agent.strip() in target_agents:
                    print(f"\n--- {filename} ({name} -> {agent}) ---")
                    sales_value = read_sales_from_xlsx(file_path)
                    print(f"最終売上値: {sales_value}")

if __name__ == "__main__":
    debug_specific_agents()