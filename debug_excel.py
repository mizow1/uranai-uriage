#!/usr/bin/env python3
"""
Excelファイルの計算エラーをデバッグするスクリプト
"""

import pandas as pd
import openpyxl
from pathlib import Path

def debug_aoki_excel():
    """202505_aoki.xlsxのM19セルとAE列の計算を確認"""
    excel_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    if not Path(excel_path).exists():
        print(f"ファイルが見つかりません: {excel_path}")
        return
    
    # openpyxlでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active
    
    print("=== 202505_aoki.xlsx デバッグ情報 ===")
    
    # M19セルの数式と値を確認
    m19_cell = ws['M19']
    print(f"M19セル:")
    print(f"  数式: {m19_cell.value}")
    
    # データ値も確認（数式と値の両方）
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    ws_data = wb_data.active
    m19_data = ws_data['M19']
    print(f"  計算結果（data_only=True）: {m19_data.value}")
    
    # 数式版も確認
    m19_formula = ws['M19']
    print(f"  計算結果（数式版）: {m19_formula.value}")
    
    # AE列の計算に関連するY列とAC列を確認
    print(f"\nAE列の計算に使用されるY列とAC列の値（23-58行）:")
    ae_manual_total = 0
    for row in range(23, 59):  # AE23:AJ58の範囲
        y_val = ws_data[f'Y{row}'].value
        ac_val = ws_data[f'AC{row}'].value
        
        if y_val is not None and ac_val is not None and y_val != "" and ac_val != "":
            try:
                calculated = float(y_val) * float(ac_val) * 0.01
                if calculated != 0:
                    print(f"  行{row}: Y{row}={y_val}, AC{row}={ac_val}, 計算値={calculated}")
                    ae_manual_total += calculated
            except (ValueError, TypeError):
                pass
    
    print(f"AE列の手動計算合計: {ae_manual_total}")
    
    # 問題の値58, 81, 163を探す
    print(f"\n問題の値58, 81, 163を検索:")
    target_values = [58, 81, 163]
    found_cells = []
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws_data.cell(row=row, column=col).value
            if cell_value in target_values:
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                found_cells.append((cell_ref, cell_value))
                print(f"  {cell_ref}: {cell_value}")
    
    if found_cells:
        print(f"合計: {sum([v for _, v in found_cells])}")
    
    # AE列の全ての値を確認
    print(f"\nAE列の全ての値:")
    ae_total = 0
    for row in range(1, ws.max_row + 1):
        cell = ws_data[f'AE{row}']
        if cell.value is not None and cell.value != 0:
            print(f"  AE{row}: {cell.value}")
            if isinstance(cell.value, (int, float)):
                ae_total += cell.value
    print(f"AE列の合計: {ae_total}")
    
    # AE列の合計を手動で計算
    ae_values = []
    for row in range(1, ws.max_row + 1):
        cell = ws_data[f'AE{row}']
        if isinstance(cell.value, (int, float)) and cell.value != 0:
            ae_values.append(cell.value)
    
    if ae_values:
        print(f"\nAE列の非ゼロ値: {ae_values}")
        print(f"AE列の合計: {sum(ae_values)}")
    
    # M19が参照している範囲を確認
    if m19_cell.value and str(m19_cell.value).startswith('='):
        print(f"\nM19セルの数式詳細: {m19_cell.value}")
        
        # SUM関数の範囲を解析
        if 'SUM(' in str(m19_cell.value):
            formula = str(m19_cell.value)
            import re
            # SUM(範囲)のパターンを抽出
            sum_pattern = r'SUM\(([^)]+)\)'
            matches = re.findall(sum_pattern, formula)
            
            for match in matches:
                print(f"  SUM範囲: {match}")
                
                # 範囲の値を確認
                if ':' in match:
                    try:
                        start_cell, end_cell = match.split(':')
                        print(f"    範囲 {match} の値:")
                        
                        # 列名と行番号を抽出
                        start_col = ''.join([c for c in start_cell if c.isalpha()])
                        start_row = int(''.join([c for c in start_cell if c.isdigit()]))
                        end_col = ''.join([c for c in end_cell if c.isalpha()])
                        end_row = int(''.join([c for c in end_cell if c.isdigit()]))
                        
                        # 範囲の値を取得（複数列対応）
                        from openpyxl.utils import column_index_from_string
                        
                        start_col_idx = column_index_from_string(start_col)
                        end_col_idx = column_index_from_string(end_col)
                        
                        total = 0
                        for col_idx in range(start_col_idx, end_col_idx + 1):
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            print(f"    列 {col_letter}:")
                            col_total = 0
                            for row in range(start_row, end_row + 1):
                                cell_ref = f"{col_letter}{row}"
                                # 数式版でも確認
                                cell_formula = ws[cell_ref].value
                                cell_value = ws_data[cell_ref].value
                                
                                if cell_value is not None and cell_value != 0:
                                    print(f"      {cell_ref}: {cell_value} (数式: {cell_formula})")
                                    if isinstance(cell_value, (int, float)):
                                        col_total += cell_value
                                        total += cell_value
                                elif cell_formula is not None and str(cell_formula).strip():
                                    print(f"      {cell_ref}: 値なし (数式: {cell_formula})")
                            if col_total > 0:
                                print(f"    列{col_letter}の合計: {col_total}")
                        print(f"    範囲全体の合計: {total}")
                    except Exception as e:
                        print(f"    範囲解析エラー: {e}")

if __name__ == '__main__':
    debug_aoki_excel()