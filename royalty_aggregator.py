#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ロイヤリティ累計推移表作成ツール

ロイヤリティフォルダ内のyyyymm形式フォルダを対象に、
各ファイルのAE19列の値を参照し、rate.csvの「名称」と「エージェント」で
マッチングして月次・累計売上を集計する。

累計は1万円以上になった翌月から0から再開。
"""

import os
import pandas as pd
import glob
from datetime import datetime
import xlwings as xw
from pathlib import Path
import re
import csv
import time
from openpyxl import load_workbook
import pywintypes
import argparse
import sys
import shutil

class RoyaltyAggregator:
    def __init__(self, royalty_dir, rate_csv_path, output_path, target_yyyymm=None, start_yyyymm=None, end_yyyymm=None):
        self.royalty_dir = royalty_dir
        self.rate_csv_path = rate_csv_path
        self.output_path = output_path
        self.target_yyyymm = target_yyyymm
        self.start_yyyymm = start_yyyymm
        self.end_yyyymm = end_yyyymm
        self.rate_data = None
        self.content_names = []
        self.monthly_data = {}
        
    def load_rate_data(self):
        """rate.csvを読み込み、コンテンツ名リストを作成"""
        try:
            self.rate_data = pd.read_csv(self.rate_csv_path, encoding='utf-8')
            print(f"Rate data loaded: {len(self.rate_data)} records")
            print(f"Columns: {list(self.rate_data.columns)}")
            
            # rate.csvの順番でコンテンツ名リストを作成
            self.content_names = self.rate_data['名称'].tolist()
            print(f"Content names: {len(self.content_names)} items")
            
            return True
        except Exception as e:
            print(f"Error loading rate data: {e}")
            return False
    
    def get_yyyymm_folders(self):
        """yyyymm形式のフォルダリストを取得"""
        folders = []
        pattern = re.compile(r'^\d{6}$')  # 6桁の数字
        
        # 期間指定の場合（開始年月と終了年月が両方指定されている場合）
        if self.start_yyyymm and self.end_yyyymm:
            if not (pattern.match(self.start_yyyymm) and pattern.match(self.end_yyyymm)):
                print(f"Invalid YYYYMM format: {self.start_yyyymm} or {self.end_yyyymm}")
                return folders
            
            # 全フォルダを取得してフィルタリング
            all_folders = []
            for item in os.listdir(self.royalty_dir):
                if pattern.match(item):
                    folder_path = os.path.join(self.royalty_dir, item)
                    if os.path.isdir(folder_path):
                        all_folders.append(item)
            
            # 期間内のフォルダを抽出
            for folder in all_folders:
                if self.start_yyyymm <= folder <= self.end_yyyymm:
                    folders.append(folder)
            
            folders.sort()  # 昇順ソート
            print(f"Period range {self.start_yyyymm} - {self.end_yyyymm}: {folders}")
            return folders
        
        # 特定年月が指定されている場合
        if self.target_yyyymm:
            if pattern.match(self.target_yyyymm):
                folder_path = os.path.join(self.royalty_dir, self.target_yyyymm)
                if os.path.isdir(folder_path):
                    folders.append(self.target_yyyymm)
                    print(f"Target YYYYMM folder: {self.target_yyyymm}")
                else:
                    print(f"Specified folder not found: {self.target_yyyymm}")
            else:
                print(f"Invalid YYYYMM format: {self.target_yyyymm}")
            return folders
        
        # 全年月を処理する場合
        for item in os.listdir(self.royalty_dir):
            if pattern.match(item):
                folder_path = os.path.join(self.royalty_dir, item)
                if os.path.isdir(folder_path):
                    folders.append(item)
        
        folders.sort()  # 昇順ソート
        print(f"Found YYYYMM folders: {folders}")
        return folders
    
    def read_sales_from_xlsx_with_openpyxl(self, file_path):
        """openpyxlを使ってxlsxファイルのAE19セルから売上データを読み取り"""
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            ae19_value = ws['AE19'].value
            wb.close()
            
            if ae19_value is not None and isinstance(ae19_value, (int, float)):
                print(f"  {os.path.basename(file_path)}: AE19={ae19_value}円 (openpyxl)")
                return int(ae19_value)
            else:
                print(f"  {os.path.basename(file_path)}: AE19に有効な値がありません (openpyxl)")
                return 0
                
        except Exception as e:
            print(f"Error reading with openpyxl {file_path}: {e}")
            return 0

    def read_sales_from_xlsx_with_xlwings(self, file_path, retry_count=0, max_retries=3):
        """xlwingsを使ってxlsxファイルのAE19セルから売上データを読み取り（リトライ機能付き）"""
        app = None
        wb = None
        
        try:
            # Excelアプリケーションを起動
            app = xw.App(visible=False, add_book=False)
            # ワークブックを開く
            wb = app.books.open(file_path)
            ws = wb.sheets[0]
            
            # AE19セルの値を取得
            ae19_value = ws.range('AE19').value
            
            # ワークブックを閉じる
            wb.close()
            # アプリケーションを終了
            app.quit()
            
            if ae19_value is not None and isinstance(ae19_value, (int, float)):
                print(f"  {os.path.basename(file_path)}: AE19={ae19_value}円 (xlwings)")
                return int(ae19_value)
            else:
                print(f"  {os.path.basename(file_path)}: AE19に有効な値がありません (xlwings)")
                return 0
            
        except pywintypes.com_error as com_err:
            # COMエラーの場合
            error_code = com_err.args[0] if com_err.args else "Unknown"
            error_msg = com_err.args[1] if len(com_err.args) > 1 else "Unknown COM error"
            
            print(f"COM Error reading {file_path}: ({error_code}) {error_msg}")
            
            # クリーンアップ
            try:
                if wb:
                    wb.close()
                if app:
                    app.quit()
            except:
                pass
            
            # リトライ判定
            if retry_count < max_retries and error_code in [-2147023170, -2147023174]:
                print(f"  Retrying ({retry_count + 1}/{max_retries})...")
                time.sleep(2)  # 2秒待機
                return self.read_sales_from_xlsx_with_xlwings(file_path, retry_count + 1, max_retries)
            else:
                print(f"  Max retries reached, switching to openpyxl...")
                return self.read_sales_from_xlsx_with_openpyxl(file_path)
                
        except Exception as e:
            print(f"General error reading {file_path}: {e}")
            # クリーンアップ
            try:
                if wb:
                    wb.close()
                if app:
                    app.quit()
            except:
                pass
            
            # 一般的なエラーの場合もopenpyxlを試す
            if retry_count == 0:
                print(f"  Switching to openpyxl...")
                return self.read_sales_from_xlsx_with_openpyxl(file_path)
            return 0

    def read_sales_from_xlsx(self, file_path):
        """xlsxファイルのAE19セルから売上データを読み取り（メインメソッド）"""
        # まずxlwingsを試し、失敗したらopenpyxlを使用
        return self.read_sales_from_xlsx_with_xlwings(file_path)
    
    def extract_name_from_filename(self, filename):
        """ファイル名から名称を抽出"""
        # ファイル名のパターン: YYYYMM_name.xlsx または name_YYYYMM.xlsx
        base_name = os.path.splitext(filename)[0]
        
        # アンダースコアで分割
        parts = base_name.split('_')
        
        # YYYYMMパターンを除去
        yyyymm_pattern = re.compile(r'^\d{6}$')
        name_parts = [part for part in parts if not yyyymm_pattern.match(part)]
        
        if name_parts:
            return name_parts[0]  # 最初の非YYYYMM部分を名称とする
        
        return base_name
    
    def process_monthly_data(self, yyyymm):
        """指定月のデータを処理"""
        folder_path = os.path.join(self.royalty_dir, yyyymm)
        if not os.path.exists(folder_path):
            print(f"Folder not exists: {folder_path}")
            return
        
        print(f"Processing {yyyymm}...")
        monthly_content_data = {}
        
        # フォルダ内のxlsxファイルを処理
        xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        
        for file_path in xlsx_files:
            filename = os.path.basename(file_path)
            name = self.extract_name_from_filename(filename)
            sales_value = self.read_sales_from_xlsx(file_path)
            
            # rate.csvにあるコンテンツのみ処理
            if name in self.content_names:
                monthly_content_data[name] = sales_value
                if sales_value > 0:
                    print(f"  {name}: {sales_value}円")
            else:
                print(f"  {name}: Not found in rate.csv")
        
        self.monthly_data[yyyymm] = monthly_content_data
        
        # 月次データ処理完了後、即座に中間保存
        self.save_intermediate_result(yyyymm)
    
    def save_intermediate_result(self, processed_yyyymm):
        """中間結果を保存"""
        try:
            # 既存ファイルがある場合はバックアップを作成
            if os.path.exists(self.output_path):
                backup_path = f"{self.output_path}.backup"
                shutil.copy2(self.output_path, backup_path)
            
            # 現在までの処理済みデータでExcelファイルを更新
            self.create_output_excel()
            print(f"  → {processed_yyyymm} processed and saved")
            
            # バックアップファイルを削除（正常保存完了後）
            backup_path = f"{self.output_path}.backup"
            if os.path.exists(backup_path):
                os.remove(backup_path)
                
        except Exception as e:
            print(f"Error saving intermediate result for {processed_yyyymm}: {e}")
            # エラーが発生した場合、バックアップファイルから復旧を試行
            backup_path = f"{self.output_path}.backup"
            if os.path.exists(backup_path):
                try:
                    shutil.copy2(backup_path, self.output_path)
                    print(f"  → Restored from backup due to save error")
                except:
                    print(f"  → Failed to restore from backup")
    
    
    def create_output_excel(self):
        """出力Excel作成"""
        sorted_months = sorted(self.monthly_data.keys())
        
        # データフレーム用のデータを準備
        data = []
        
        for month in sorted_months:
            row = {'売上年月': month}
            
            # rate.csvの順番でコンテンツを処理
            for content_name in self.content_names:
                sales_value = self.monthly_data.get(month, {}).get(content_name, 0)
                row[content_name] = sales_value
            
            data.append(row)
        
        # DataFrameを作成
        df = pd.DataFrame(data)
        
        # Excelファイルに出力
        df.to_excel(self.output_path, index=False, engine='openpyxl')
        print(f"Output created: {self.output_path}")
        
        # CSV版も作成
        csv_path = self.output_path.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        print(f"CSV output created: {csv_path}")
    
    def run(self):
        """メイン処理実行"""
        print("=== ロイヤリティ累計推移表作成開始 ===")
        
        # rate.csvの読み込み
        if not self.load_rate_data():
            return False
        
        # yyyymmフォルダの取得
        yyyymm_folders = self.get_yyyymm_folders()
        if not yyyymm_folders:
            print("No YYYYMM folders found")
            return False
        
        # 各月のデータを処理（各月毎に保存される）
        for yyyymm in yyyymm_folders:
            try:
                self.process_monthly_data(yyyymm)
            except Exception as e:
                print(f"Error processing {yyyymm}: {e}")
                print("Continuing with next month...")
                continue
        
        # 最終確認として再保存（万が一のため）
        if self.monthly_data:
            print("Creating final output...")
            self.create_output_excel()
        
        print("=== 処理完了 ===")
        return True


def main():
    # コマンドライン引数の解析
    parser = argparse.ArgumentParser(description='ロイヤリティ累計推移表作成ツール')
    parser.add_argument('yyyymm', nargs='*', help='処理対象年月 (YYYYMM形式)\n  省略時: 全年月を処理\n  1つ指定: 指定月のみ処理\n  2つ指定: 指定期間を処理 (例: 202401 202403)')
    args = parser.parse_args()
    
    # パスの設定
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    rate_csv_path = r"C:\Users\OW\pj\uriage\rate.csv"
    output_path = r"C:\Users\OW\pj\uriage\コンテンツ別累計推移表.xlsx"
    
    # 引数の処理と出力ファイル名の設定
    target_yyyymm = None
    start_yyyymm = None
    end_yyyymm = None
    
    if len(args.yyyymm) == 1:
        # 単一年月指定
        target_yyyymm = args.yyyymm[0]
        base_name = os.path.splitext(output_path)[0]
        extension = os.path.splitext(output_path)[1]
        output_path = f"{base_name}_{target_yyyymm}{extension}"
        print(f"指定年月: {target_yyyymm}")
    elif len(args.yyyymm) == 2:
        # 期間指定
        start_yyyymm = args.yyyymm[0]
        end_yyyymm = args.yyyymm[1]
        # 開始年月と終了年月の順序を確認
        if start_yyyymm > end_yyyymm:
            start_yyyymm, end_yyyymm = end_yyyymm, start_yyyymm
        base_name = os.path.splitext(output_path)[0]
        extension = os.path.splitext(output_path)[1]
        output_path = f"{base_name}_{start_yyyymm}-{end_yyyymm}{extension}"
        print(f"指定期間: {start_yyyymm} ～ {end_yyyymm}")
    elif len(args.yyyymm) > 2:
        print("エラー: 年月は最大2つまで指定できます")
        sys.exit(1)
    else:
        print("全年月を処理します")
    
    # 集計実行
    aggregator = RoyaltyAggregator(royalty_dir, rate_csv_path, output_path, target_yyyymm, start_yyyymm, end_yyyymm)
    success = aggregator.run()
    
    if success:
        print("\n集計処理が正常に完了しました。")
    else:
        print("\n集計処理でエラーが発生しました。")


if __name__ == "__main__":
    main()