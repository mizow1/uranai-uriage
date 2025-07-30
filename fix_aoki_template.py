#!/usr/bin/env python3
"""
aoki.xlsxテンプレートのAE列の数式を修正するスクリプト
"""

import openpyxl
from pathlib import Path
import shutil

def fix_aoki_template():
    """aoki.xlsxテンプレートのAE列数式を修正"""
    template_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\コンテンツ関連支払明細書フォーマット\aoki.xlsx"
    backup_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\コンテンツ関連支払明細書フォーマット\bk\aoki-bk.xlsx"
    
    if not Path(template_path).exists():
        print(f"テンプレートファイルが見つかりません: {template_path}")
        return
    
    # バックアップを作成
    backup_dir = Path(backup_path).parent
    backup_dir.mkdir(exist_ok=True)
    shutil.copy2(template_path, backup_path)
    print(f"バックアップを作成: {backup_path}")
    
    # テンプレートファイルを開く
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    print("=== 現在の数式 ===")
    
    # 修正対象の行範囲（23-58行）
    for row in range(23, 59):
        cell_ref = f"AE{row}"
        cell = ws[cell_ref]
        
        if cell.value and str(cell.value).startswith('='):
            old_formula = cell.value
            print(f"{cell_ref}: {old_formula}")
            
            # 数式を修正：ROUND関数で整数に丸める
            # 元: =IF(OR(Y23="", AC23=""), "", IF(Y23*AC23*0.01=0, "", Y23*AC23*0.01))
            # 新: =IF(OR(Y23="", AC23=""), "", IF(Y23*AC23*0.01=0, "", ROUND(Y23*AC23*0.01,0)))
            
            new_formula = old_formula.replace(
                "Y{0}*AC{0}*0.01))".format(row),
                "ROUND(Y{0}*AC{0}*0.01,0)))".format(row)
            )
            
            # 最後の部分だけが置換されない場合の対処
            if new_formula == old_formula:
                # より汎用的な置換
                import re
                pattern = r'(Y\d+\*AC\d+\*0\.01)\)\)$'
                replacement = r'ROUND(\1,0)))'
                new_formula = re.sub(pattern, replacement, old_formula)
            
            if new_formula != old_formula:
                ws[cell_ref].value = new_formula
                print(f"  → {new_formula}")
            else:
                print(f"  変更なし")
    
    # ファイルを保存
    wb.save(template_path)
    print(f"\nテンプレートファイルを更新しました: {template_path}")
    
    # 修正後の確認
    print("\n=== 修正後の確認 ===")
    wb_check = openpyxl.load_workbook(template_path)
    ws_check = wb_check.active
    
    for row in [23, 24, 25, 26, 27]:
        cell_ref = f"AE{row}"
        cell = ws_check[cell_ref]
        if cell.value and str(cell.value).startswith('='):
            print(f"{cell_ref}: {cell.value}")

if __name__ == '__main__':
    fix_aoki_template()