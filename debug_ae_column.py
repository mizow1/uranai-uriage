#\!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AE19セルの問題をデバッグするためのスクリプト
"""

import openpyxl
import glob
import os

def debug_xlsx_file(file_path):
    """xlsxファイルのセル内容をデバッグ"""
    print(f"=== {os.path.basename(file_path)} の分析 ===")
    
    try:
        # 数式ブックと計算値ブックの両方を読み込み
        workbook_formula = openpyxl.load_workbook(file_path, data_only=False)
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        
        worksheet_formula = workbook_formula.active
        worksheet_data = workbook_data.active
        
        print(f"シート名: {worksheet_formula.title}")
        print(f"最大行: {worksheet_formula.max_row}, 最大列: {worksheet_formula.max_column}")
        
        # AE19セルの詳細分析
        ae19_formula = worksheet_formula.cell(row=19, column=31)
        ae19_data = worksheet_data.cell(row=19, column=31)
        
        print(f"\nAE19セル分析:")
        print(f"  数式: {ae19_formula.value}")
        print(f"  データタイプ: {type(ae19_formula.value)}")
        print(f"  計算値: {ae19_data.value}")
        print(f"  計算値タイプ: {type(ae19_data.value)}")
        
        # M19, S19, Y19セルの分析
        print(f"\n参照セル分析:")
        for col_name, col_num in [("M", 13), ("S", 19), ("Y", 25)]:
            cell_formula = worksheet_formula.cell(row=19, column=col_num)
            cell_data = worksheet_data.cell(row=19, column=col_num)
            print(f"  {col_name}19 - 数式: {cell_formula.value}, 計算値: {cell_data.value}")
        
        # 19行目の周辺セルをチェック（M～Y列）
        print(f"\n19行目のM～Y列データ:")
        for col in range(13, 26):  # M列(13)からY列(25)まで
            cell_formula = worksheet_formula.cell(row=19, column=col)
            cell_data = worksheet_data.cell(row=19, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            if cell_data.value not in [None, 0, ""]:
                print(f"  {col_letter}19: 数式={cell_formula.value}, 値={cell_data.value}")
        
        # 他の行で有効な値があるかチェック
        print(f"\nAE列の他の行（非0値のみ）:")
        for row in range(1, 61):  # 1-60行目
            cell_data = worksheet_data.cell(row=row, column=31)  # AE列
            if cell_data.value not in [None, 0, ""]:
                cell_formula = worksheet_formula.cell(row=row, column=31)
                print(f"  AE{row}: 数式={cell_formula.value}, 値={cell_data.value}")
        
        workbook_formula.close()
        workbook_data.close()
        
    except Exception as e:
        print(f"エラー: {e}")

def main():
    # テスト用に1つのファイルを分析
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    
    # 202508フォルダから1つのファイルを選択
    test_folder = os.path.join(royalty_dir, "202508")
    if os.path.exists(test_folder):
        xlsx_files = glob.glob(os.path.join(test_folder, "*.xlsx"))
        if xlsx_files:
            # 最初のファイルをテスト
            debug_xlsx_file(xlsx_files[0])
            
            # 複数ファイルもテスト
            if len(xlsx_files) > 1:
                print(f"\n" + "="*50)
                debug_xlsx_file(xlsx_files[1])
        else:
            print("No xlsx files found")
    else:
        print("Test folder not found")

if __name__ == "__main__":
    main()
