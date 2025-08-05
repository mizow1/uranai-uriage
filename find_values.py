#\!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
より詳細なExcelファイルの分析
"""

import openpyxl
import glob
import os

def find_all_values_in_xlsx(file_path):
    """xlsxファイル内のすべての数値を検索"""
    print(f"=== {os.path.basename(file_path)} の全体分析 ===")
    
    try:
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        worksheet_data = workbook_data.active
        
        print(f"シート名: {worksheet_data.title}")
        print(f"最大行: {worksheet_data.max_row}, 最大列: {worksheet_data.max_column}")
        
        # すべてのセルをスキャンして数値を探す
        print(f"\n数値が含まれるセル（0以外）:")
        found_values = []
        
        for row in range(1, min(worksheet_data.max_row + 1, 100)):  # 最初の100行をチェック
            for col in range(1, min(worksheet_data.max_column + 1, 50)):  # 最初の50列をチェック
                cell_value = worksheet_data.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)) and cell_value != 0:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    found_values.append((f"{col_letter}{row}", cell_value))
        
        if found_values:
            print(f"  見つかった数値: {len(found_values)}個")
            # 最初の20個を表示
            for cell_ref, value in found_values[:20]:
                print(f"    {cell_ref}: {value}")
            if len(found_values) > 20:
                print(f"    ... (他に{len(found_values) - 20}個)")
        else:
            print("  数値が見つかりませんでした")
        
        # AE列のすべての行をチェック
        print(f"\nAE列の全データ:")
        ae_values = []
        for row in range(1, min(worksheet_data.max_row + 1, 100)):
            cell_value = worksheet_data.cell(row=row, column=31).value  # AE列 = 31列目
            if cell_value is not None and cell_value != "":
                ae_values.append((row, cell_value))
        
        if ae_values:
            for row, value in ae_values:
                print(f"  AE{row}: {value} (type: {type(value)})")
        else:
            print("  AE列にデータが見つかりませんでした")
        
        workbook_data.close()
        
    except Exception as e:
        print(f"エラー: {e}")

def main():
    # テスト用に1つのファイルを分析
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    
    # 202508フォルダから1つのファイルを選択
    test_folder = os.path.join(royalty_dir, "202508")
    if os.path.exists(test_folder):
        xlsx_files = glob.glob(os.path.join(test_folder, "*.xlsx"))
        if xlsx_files:
            # 最初のファイルをテスト
            find_all_values_in_xlsx(xlsx_files[0])
        else:
            print("No xlsx files found")
    else:
        print("Test folder not found")

if __name__ == "__main__":
    main()
