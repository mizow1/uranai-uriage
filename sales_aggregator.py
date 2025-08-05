import os
import pandas as pd
import openpyxl
from pathlib import Path
import re
import csv
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# 共通コンポーネントをインポート
from common import (
    CSVHandler,
    ExcelHandler, 
    UnifiedLogger,
    ErrorHandler,
    ConfigManager,
    ProcessingResult,
    ContentDetail,
    ProcessingSummary
)

class SalesAggregator:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.results = []
        # 共通コンポーネントを初期化
        self.logger = UnifiedLogger(__name__, log_file=Path("logs/sales_aggregator.log"))
        self.error_handler = ErrorHandler(self.logger.logger)
        self.csv_handler = CSVHandler(self.logger.logger, self.error_handler)
        self.excel_handler = ExcelHandler(self.logger.logger, self.error_handler)
        self.config = ConfigManager(logger=self.logger.logger)
        
        # docomoのKEIKOソウルメイト占術の統一コンテンツ名を定義
        self.DOCOMO_KEIKO_UNIFIED_NAME = "ＫＥＩＫＯ☆ソウルメイト占術（統合）"
    
    def _load_encrypted_workbook(self, file_path: Path, passwords: list):
        """暗号化されたワークブックを読み込み"""
        try:
            import msoffcrypto
            import io
            
            for password in passwords:
                try:
                    with open(file_path, 'rb') as file:
                        office_file = msoffcrypto.OfficeFile(file)
                        office_file.load_key(password=password if password else None)
                        
                        decrypted = io.BytesIO()
                        office_file.save(decrypted)
                        decrypted.seek(0)
                        
                        wb = openpyxl.load_workbook(decrypted, data_only=True)
                        self.logger.info(f"パスワード解除成功: {file_path.name} ('{password}')")
                        return wb
                        
                except Exception:
                    continue
            
            return None
            
        except ImportError:
            self.logger.error("msoffcrypto-toolが必要です")
            return None
        
    def find_files_in_yearmonth_folders(self):
        """年月フォルダ内のファイルを検索"""
        files_by_platform = {
            'ameba': [],
            'rakuten': [],
            'mediba': [],
            'excite': [],
            'line': [],
            'docomo': [],
            'au': [],
            'softbank': []
        }
        
        # 年フォルダ（4桁）内の月フォルダ（6桁）を検索
        for year_folder in self.base_path.iterdir():
            if year_folder.is_dir() and re.match(r'\d{4}', year_folder.name):
                for month_folder in year_folder.iterdir():
                    if month_folder.is_dir() and re.match(r'\d{6}', month_folder.name):
                        # 月フォルダ直下のファイルを検索
                        for file in month_folder.iterdir():
                            if file.is_file():
                                filename = file.name.lower()
                                if '【株式会社アウトワード御中】satori実績_' in file.name or 'satori実績_' in filename:
                                    files_by_platform['ameba'].append(file)
                                elif 'rcms' in filename:
                                    files_by_platform['rakuten'].append(file)
                                elif 'salessummary' in filename:
                                    files_by_platform['mediba'].append(file)
                                elif 'excite' in filename:
                                    files_by_platform['excite'].append(file)
                                elif 'bp40000746' in filename and filename.endswith('.csv'):
                                    files_by_platform['docomo'].append(file)
                                elif 'cp02お支払い明細書' in filename and (filename.endswith('.pdf') or filename.endswith('.csv')):
                                    files_by_platform['au'].append(file)
                                elif 'oid_pay_9ati' in filename and filename.endswith('.pdf'):
                                    files_by_platform['softbank'].append(file)
                            # サブフォルダも検索（LINEファイル、softbankファイル、auファイル用）
                            elif file.is_dir():
                                for subfile in file.iterdir():
                                    if subfile.is_file():
                                        subfilename = subfile.name.lower()
                                        if subfilename.startswith('line-contents-') and (subfile.suffix.lower() in ['.xls', '.xlsx', '.csv']):
                                            files_by_platform['line'].append(subfile)
                                        elif 'oid_pay_9ati' in subfilename and subfile.suffix.lower() == '.pdf':
                                            files_by_platform['softbank'].append(subfile)
                                        elif 'cp02お支払い明細書' in subfile.name and (subfile.suffix.lower() in ['.pdf', '.csv']):
                                            files_by_platform['au'].append(subfile)
                                    # さらに深いサブフォルダも検索（softbankファイル、auファイル用）
                                    elif subfile.is_dir():
                                        for subsubfile in subfile.iterdir():
                                            if subsubfile.is_file():
                                                subsubfilename = subsubfile.name.lower()
                                                if 'oid_pay_9ati' in subsubfilename and subsubfile.suffix.lower() == '.pdf':
                                                    files_by_platform['softbank'].append(subsubfile)
                                                elif 'cp02お支払い明細書' in subsubfile.name and (subsubfile.suffix.lower() in ['.pdf', '.csv']):
                                                    files_by_platform['au'].append(subsubfile)
        
        return files_by_platform
    
    def process_ameba_file(self, file_path: Path) -> ProcessingResult:
        """ameba占い（SATORI実績）ファイルを処理"""
        result = ProcessingResult(
            platform="ameba",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # 統一Excelハンドラーを使用してファイルを読み込み
            passwords = self.config.get_processing_settings().get('excel_passwords', ['', 'password', '123456', '000000', 'admin', 'user'])
            
            # ExcelHandlerでパスワード保護を処理
            df = self.excel_handler.read_excel_with_password_handling(file_path, passwords=passwords)
            
            if df is None:
                # 直接openpyxlで試行（シート単位での処理が必要なため）
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as e:
                    if "password" in str(e).lower() or "protected" in str(e).lower():
                        wb = self._load_encrypted_workbook(file_path, passwords)
                        if wb is None:
                            result.add_error("パスワード保護解除に失敗")
                            return result
                    else:
                        result.add_error(f"ファイル読み込みエラー: {str(e)}")
                        return result
            else:
                # DataFrameからワークブックを取得（複雑な処理のため直接openpyxl使用）
                wb = openpyxl.load_workbook(file_path, data_only=True)
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # 新仕様：各シートから情報提供料を集計し、同一コンテンツは合算する
            content_groups = {}
            content_counts = {}  # 件数を管理
            
            # 1. 「従量実績」シートでC列の値が一致するもののJ列の合計額を算出
            self._process_sheet_data(wb, '従量実績', content_groups, content_counts, 2, 9)  # C列, J列
            
            # 2. 「docomo占い」シートでC列の値が一致するもののJ列の合計額を算出
            self._process_sheet_data(wb, 'docomo占い', content_groups, content_counts, 2, 9)  # C列, J列
            
            # 3. 「月額実績」シートのB列の値が一致するもののE列の合計額を算出
            self._process_sheet_data(wb, '月額実績', content_groups, content_counts, 1, 4)  # B列, E列
            
            # 合計値を計算
            情報提供料合計 = sum(content_groups.values())
            実績合計 = 情報提供料合計 / 0.3 if 情報提供料合計 > 0 else 0  # 30%を除算した値
            
            # ContentDetailリストを作成
            for content_name, 情報提供料 in content_groups.items():
                実績 = 情報提供料 / 0.3 if 情報提供料 > 0 else 0
                件数 = content_counts.get(content_name, 0)
                detail = ContentDetail(
                    content_group=str(content_name),
                    performance=round(実績),
                    information_fee=round(情報提供料),
                    sales_count=件数
                )
                result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'content_groups_count': len(content_groups),
                '情報提供料': round(情報提供料合計),
                '実績合計': round(実績合計)
            }
            
            self.logger.info(f"ameba処理完了: {len(content_groups)}コンテンツグループ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def _process_sheet_data(self, wb, sheet_name: str, content_groups: dict, content_counts: dict, key_col: int, value_col: int):
        """シートデータを処理してコンテンツグループに集計"""
        try:
            if sheet_name not in wb.sheetnames:
                self.logger.warning(f"シート '{sheet_name}' が存在しません")
                return
                
            sheet = wb[sheet_name]
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df.drop(0).reset_index(drop=True)
            
            for _, row in df.iterrows():
                key_value = row.iloc[key_col] 
                amount_value = row.iloc[value_col]
                
                if pd.notna(key_value) and pd.notna(amount_value):
                    amount_numeric = pd.to_numeric(amount_value, errors='coerce')
                    if pd.notna(amount_numeric):
                        if key_value not in content_groups:
                            content_groups[key_value] = 0
                            content_counts[key_value] = 0
                        content_groups[key_value] += amount_numeric
                        
                        # 「従量実績」「docomo占い」シートでのみF列の値を件数として使用
                        if sheet_name in ['従量実績', 'docomo占い']:
                            count_value = row.iloc[5] if len(row) > 5 else 1
                            count_numeric = pd.to_numeric(count_value, errors='coerce')
                            if pd.notna(count_numeric):
                                content_counts[key_value] += count_numeric
                            else:
                                content_counts[key_value] += 1  # F列が無効な場合は1を加算
                        # 「月額実績」シートでは件数算出に使用しない（何もしない）
                        
        except Exception as e:
            self.logger.warning(f"{sheet_name}シート処理エラー: {e}")
    
    def process_rakuten_file(self, file_path: Path) -> ProcessingResult:
        """楽天占い（rcms・楽天明細）ファイルを処理"""
        result = ProcessingResult(
            platform="rakuten",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # ファイル名にrcmsを含むファイルのみ処理
            if 'rcms' not in file_path.name.lower():
                result.add_error("rcmsファイルではありません")
                self.logger.warning(f"楽天占いファイル処理スキップ: {file_path.name} - rcmsファイルではありません")
                return result
            
            # ファイル拡張子に応じてデータを読み込み
            if file_path.suffix.lower() == '.csv':
                df = self.csv_handler.read_csv_with_encoding_detection(file_path)
            else:
                df = self.excel_handler.read_excel_with_password_handling(file_path)
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # RCMSファイルの処理
            # L列の値「hoge_xxx」のhoge部分が一致するもののN列の値で計算
            if len(df.columns) < 14:
                result.add_error(f"列数が不足: 必要14列以上、実際{len(df.columns)}列")
                return result
                
            l_column = df.iloc[:, 11]  # L列
            n_column = df.iloc[:, 13]  # N列
            
            hoge_groups = {}
            for i, value in enumerate(l_column):
                if pd.notna(value) and '_' in str(value):
                    hoge_part = str(value).split('_')[0]
                    if hoge_part not in hoge_groups:
                        hoge_groups[hoge_part] = []
                    hoge_groups[hoge_part].append(n_column.iloc[i])
            
            # 各グループの計算
            for hoge, values in hoge_groups.items():
                group_sum = sum(pd.to_numeric(v, errors='coerce') for v in values if pd.notna(v))
                実績_sum = group_sum / 1.1  # N列の値の合計額を1.1で除算
                情報提供料_sum = 実績_sum * 0.725  # 実績に0.725を乗算
                
                detail = ContentDetail(
                    content_group=hoge,
                    performance=round(実績_sum),
                    information_fee=round(情報提供料_sum),
                    sales_count=len(values)  # 件数を追加
                )
                result.add_detail(detail)
            
            # 「月額実績」シートの処理も追加（もしあれば）
            if file_path.suffix.lower() in ['.xlsx', '.xls']:
                try:
                    monthly_df = self.excel_handler.read_excel_with_password_handling(
                        file_path, sheet_name='月額実績'
                    )
                    if monthly_df is not None and len(monthly_df.columns) >= 5:
                        monthly_amount = monthly_df.iloc[:, 4].sum()  # E列
                        monthly_実績 = monthly_amount / 1.1
                        monthly_情報提供料 = monthly_実績 * 0.725
                        
                        detail = ContentDetail(
                            content_group='月額実績',
                            performance=round(monthly_実績),
                            information_fee=round(monthly_情報提供料),
                            sales_count=len(monthly_df)  # 件数を追加
                        )
                        result.add_detail(detail)
                except Exception as e:
                    self.logger.warning(f"月額実績シート処理エラー: {str(e)}")
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'hoge_groups_count': len(hoge_groups),
                'rcms_processing': True
            }
            
            self.logger.info(f"楽天処理完了: {len(hoge_groups)}グループ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_au_file(self, file_path: Path) -> ProcessingResult:
        """mediba占い（SalesSummary）ファイルを処理"""
        result = ProcessingResult(
            platform="mediba",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # ファイル拡張子に応じてデータを読み込み
            if file_path.suffix.lower() == '.csv':
                df = self.csv_handler.read_csv_with_encoding_detection(file_path)
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                df = self.excel_handler.read_excel_with_password_handling(file_path)
                if df is None:
                    # 複数エンジンで試行
                    df = self.excel_handler.try_multiple_engines(file_path)
            else:
                result.add_error(f"サポートされていないファイル形式: {file_path.suffix}")
                return result
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # プラットフォーム名決定（mediba占いとして統一処理）
            platform_name = "mediba"
            
            # 列数チェック
            if len(df.columns) < 11:
                result.add_error(f"列数が不足: 必要11列以上、実際{len(df.columns)}列")
                return result
            
            # B列（インデックス1）でグループ化してG列の合計を計算
            b_column = df.iloc[:, 1]  # B列
            g_column = df.iloc[:, 6]  # G列
            k_column = df.iloc[:, 10]  # K列
            
            # 数値に変換
            g_column = pd.to_numeric(g_column, errors='coerce').fillna(0)
            k_column = pd.to_numeric(k_column, errors='coerce').fillna(0)
            
            b_groups = {}
            for i, b_value in enumerate(b_column):
                if pd.notna(b_value):
                    if b_value not in b_groups:
                        b_groups[b_value] = {'g_values': [], 'k_values': []}
                    b_groups[b_value]['g_values'].append(g_column.iloc[i])
                    b_groups[b_value]['k_values'].append(k_column.iloc[i])
            
            # 各グループの計算
            for b_value, values in b_groups.items():
                g_sum = sum(values['g_values'])
                k_sum = sum(values['k_values'])
                実績_sum = g_sum  # G列の値
                情報提供料_sum = (g_sum * 0.4) - k_sum  # G列の40%からK列を引いた値
                
                detail = ContentDetail(
                    content_group=str(b_value),
                    performance=round(実績_sum),
                    information_fee=round(情報提供料_sum),
                    sales_count=len(values['g_values'])  # 件数を追加
                )
                result.add_detail(detail)
            
            # 合計を計算  
            result.calculate_totals()
            result.success = True
            result.platform = platform_name  # プラットフォーム名を上書き
            result.metadata = {
                'b_groups_count': len(b_groups),
                'platform_name': platform_name
            }
            
            self.logger.info(f"mediba処理完了: {len(b_groups)}グループ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_excite_file(self, file_path: Path) -> ProcessingResult:
        """excite占いファイルを処理"""
        result = ProcessingResult(
            platform="excite",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # ファイル拡張子に応じてデータを読み込み
            if file_path.suffix.lower() == '.csv':
                # CSVファイルの場合、特殊な読み込み処理（説明文をスキップ）
                df = self.csv_handler.read_csv_with_encoding_detection(
                    file_path, skiprows=3, on_bad_lines='skip', engine='python'
                )
            else:
                df = self.excel_handler.read_excel_with_password_handling(file_path)
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # データの妥当性チェック
            if df.shape[1] < 5 or df.shape[0] == 0:
                result.add_error("データが不正または空です")
                return result
            
            # 数値列の存在チェック
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) == 0:
                result.add_error("数値列が見つかりません")
                return result
            
            # exciteファイル固有の処理ロジック - コンテンツ別処理に変更
            # コンテンツ名列を検索（一般的な列名を検索）
            content_name_col = None
            possible_content_names = ['コンテンツ名', 'コンテンツ', 'サービス名', 'サービス', 'アプリ名', 'アプリ', '商品名', '商品']
            
            for col_name in df.columns:
                col_name_str = str(col_name).strip()
                if any(keyword in col_name_str for keyword in possible_content_names):
                    content_name_col = col_name
                    break
            
            if content_name_col is None:
                # コンテンツ名列が見つからない場合は、最初の非数値列を使用
                text_cols = df.select_dtypes(include=['object', 'string']).columns
                if len(text_cols) > 0:
                    content_name_col = text_cols[0]
                    self.logger.warning(f"コンテンツ名列が特定できないため、{content_name_col}列を使用します")
                else:
                    # それでも見つからない場合は従来通り一括処理（exciteの場合はF列のみ）
                    self.logger.warning("コンテンツ名列が見つからないため、一括処理を行います")
                    total_amount = 0
                    if len(df.columns) > 5:
                        # exciteの場合はF列（インデックス5）のみを使用
                        total_amount = df.iloc[:, 5].sum()
                    else:
                        # F列がない場合は従来通り
                        for col in numeric_cols:
                            column_sum = df[col].sum()
                            if column_sum > 0:
                                total_amount += column_sum
                    
                    information_fee = total_amount * 0.6  # exciteは60%が情報提供料
                    
                    detail = ContentDetail(
                        content_group="excite_total",
                        performance=round(total_amount),
                        information_fee=round(information_fee),
                        sales_count=len(df)  # 件数を追加
                    )
                    result.add_detail(detail)
            else:
                # コンテンツ名列が見つかった場合、コンテンツ別に処理
                self.logger.info(f"exciteコンテンツ名列として {content_name_col} を使用")
                
                # コンテンツ名でグループ化
                content_groups = {}
                
                for _, row in df.iterrows():
                    content_name = row[content_name_col]
                    if pd.notna(content_name) and str(content_name).strip():
                        content_name = str(content_name).strip()
                        if content_name not in content_groups:
                            content_groups[content_name] = []
                        
                        # exciteの場合はF列（金額列、インデックス5）のみを使用
                        row_total = 0
                        if len(df.columns) > 5:
                            f_col_value = pd.to_numeric(row.iloc[5], errors='coerce')
                            if pd.notna(f_col_value) and f_col_value > 0:
                                row_total = f_col_value
                        
                        if row_total > 0:
                            content_groups[content_name].append(row_total)
                
                # 各コンテンツの計算結果を追加
                for content_name, amounts in content_groups.items():
                    if amounts:  # 空でない場合のみ処理
                        total_amount = sum(amounts)
                        information_fee = total_amount * 0.6  # exciteは60%が情報提供料
                        
                        detail = ContentDetail(
                            content_group=content_name,
                            performance=round(total_amount),
                            information_fee=round(information_fee),
                            sales_count=len(amounts)  # 件数を追加
                        )
                        result.add_detail(detail)
                
                if not content_groups:
                    # コンテンツが見つからない場合はエラー
                    result.add_error("有効なコンテンツデータが見つかりません")
                    return result
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'numeric_columns': len(numeric_cols),
                'data_rows': len(df)
            }
            
            self.logger.info(f"excite処理完了: {len(df)}行処理")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_line_file(self, file_path: Path) -> ProcessingResult:
        """LINE占いファイルを処理 - 新仕様：「内訳」シートから「アイテム名」「RS対象額」「RS金額」を処理"""
        result = ProcessingResult(
            platform="line",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # ファイル名にLINEを含むファイル（xls, xlsx, csv）を処理
            if 'line' not in file_path.name.lower() or file_path.suffix.lower() not in ['.xls', '.xlsx', '.csv']:
                result.add_error("LINEを含むファイルではありません")
                self.logger.warning(f"LINE占いファイル処理スキップ: {file_path.name}")
                return result
            
            # 既に集計済みのcontentsファイルかどうかをチェック
            if 'line-contents-' in file_path.name.lower():
                self.logger.info(f"集計済みLINEファイルを検出: {file_path.name} - 集計済みデータとして処理")
                # 集計済みファイルの場合は直接CSVデータを読み込む
                return self._process_aggregated_line_file(file_path)
            
            # ファイル形式に応じて読み込み処理を分岐
            if file_path.suffix.lower() == '.csv':
                # CSVファイルの場合
                df = self.csv_handler.read_csv_with_encoding_detection(file_path)
                if df is None or df.empty:
                    result.add_error("CSVファイルの読み込みに失敗")
                    return result
                
                self.logger.log_file_operation("読み込み", file_path, True)
                
                # CSVの場合は直接DataFrameとして処理
                if len(df) < 2:
                    result.add_error("CSVファイルにデータがありません")
                    return result
                
                # CSVファイル用の処理（既に集計済みデータの場合）
                if 'コンテンツ名' in df.columns and '実績' in df.columns and '情報提供料' in df.columns:
                    # 既に集計済みの形式の場合は直接読み込み
                    for _, row in df.iterrows():
                        content_name = row['コンテンツ名']
                        performance = pd.to_numeric(row['実績'], errors='coerce')
                        info_fee = pd.to_numeric(row['情報提供料'], errors='coerce')
                        sales_count = pd.to_numeric(row.get('売上件数', 0), errors='coerce') if '売上件数' in row else 0
                        
                        if pd.notna(content_name) and (performance > 0 or info_fee > 0):
                            detail = ContentDetail(
                                content_group=str(content_name),
                                performance=round(performance) if pd.notna(performance) else 0,
                                information_fee=round(info_fee) if pd.notna(info_fee) else 0,
                                sales_count=round(sales_count) if pd.notna(sales_count) else 0
                            )
                            result.add_detail(detail)
                    
                    # 合計を計算
                    result.calculate_totals()
                    result.success = True
                    result.metadata = {
                        'content_count': len(result.details),
                        'csv_format': '集計済み'
                    }
                    
                    self.logger.info(f"LINE CSV処理完了（集計済み形式）: {len(result.details)}コンテンツ")
                    return result
                
            else:
                # Excelファイルの場合（従来の処理）
                passwords = self.config.get_processing_settings().get('excel_passwords', ['', 'password', '123456', '000000', 'admin', 'user'])
                
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as e:
                    if "password" in str(e).lower() or "protected" in str(e).lower():
                        wb = self._load_encrypted_workbook(file_path, passwords)
                        if wb is None:
                            result.add_error("パスワード保護解除に失敗")
                            return result
                    else:
                        result.add_error(f"ファイル読み込みエラー: {str(e)}")
                        return result
                
                self.logger.log_file_operation("読み込み", file_path, True)
                
                # 「内訳」シートを検索
                breakdown_sheet = None
                for sheet_name in wb.sheetnames:
                    if '内訳' in sheet_name:
                        breakdown_sheet = wb[sheet_name]
                        break
                
                if breakdown_sheet is None:
                    result.add_error("「内訳」シートが見つかりません")
                    return result
                
                # DataFrameに変換
                df = pd.DataFrame(breakdown_sheet.values)
                if df.empty or len(df) < 2:
                    result.add_error("「内訳」シートにデータがありません")
                    return result
            
            # 1行目をヘッダーとして使用
            headers = df.iloc[0].astype(str).tolist()
            df.columns = headers
            df = df.drop(0).reset_index(drop=True)
            
            # 必要な列を検索
            item_name_col = None
            rs_target_col = None
            rs_amount_col = None
            
            for i, header in enumerate(headers):
                if 'アイテム名' in str(header):
                    item_name_col = i
                elif 'RS対象額' in str(header):
                    rs_target_col = i
                elif 'RS金額' in str(header):
                    rs_amount_col = i
            
            if item_name_col is None:
                result.add_error("「アイテム名」列が見つかりません")
                return result
            if rs_target_col is None:
                result.add_error("「RS対象額」列が見つかりません")
                return result
            if rs_amount_col is None:
                result.add_error("「RS金額」列が見つかりません")
                return result
            
            # アイテム名でグループ化して集計
            item_groups = {}
            
            for _, row in df.iterrows():
                item_name = row.iloc[item_name_col]
                rs_target = row.iloc[rs_target_col]
                rs_amount = row.iloc[rs_amount_col]
                
                if pd.notna(item_name):
                    item_name = str(item_name).strip()
                    if item_name not in item_groups:
                        item_groups[item_name] = {'rs_target': 0, 'rs_amount': 0, 'count': 0}
                    
                    # RS対象額を加算
                    if pd.notna(rs_target):
                        rs_target_numeric = pd.to_numeric(rs_target, errors='coerce')
                        if pd.notna(rs_target_numeric):
                            item_groups[item_name]['rs_target'] += rs_target_numeric
                    
                    # RS金額を加算
                    if pd.notna(rs_amount):
                        rs_amount_numeric = pd.to_numeric(rs_amount, errors='coerce')
                        if pd.notna(rs_amount_numeric):
                            item_groups[item_name]['rs_amount'] += rs_amount_numeric
                    
                    # 件数をカウント
                    item_groups[item_name]['count'] += 1
            
            # 各アイテムの計算
            for item_name, values in item_groups.items():
                # RS対象額の合計を1.1で除算→「実績」
                実績 = values['rs_target'] / 1.1 if values['rs_target'] > 0 else 0
                # RS金額の合計を1.1で除算→「情報提供料」
                情報提供料 = values['rs_amount'] / 1.1 if values['rs_amount'] > 0 else 0
                # 実際の件数を使用
                件数 = values['count']
                
                detail = ContentDetail(
                    content_group=item_name,
                    performance=round(実績),
                    information_fee=round(情報提供料),
                    sales_count=件数
                )
                result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'item_groups_count': len(item_groups),
                'sheet_used': '内訳'
            }
            
            self.logger.info(f"LINE処理完了（新仕様）: {len(item_groups)}アイテムグループ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def _process_aggregated_line_file(self, file_path: Path) -> ProcessingResult:
        """集計済みLINEファイル（line-contents-）を処理"""
        result = ProcessingResult(
            platform="line",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # CSVファイルを読み込み
            df = self.csv_handler.read_csv_with_encoding_detection(file_path)
            if df is None or df.empty:
                result.add_error("集計済みCSVファイルの読み込みに失敗")
                return result
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # 集計済みファイルの列構成をチェック
            expected_columns = ['コンテンツ名', '実績', '情報提供料']
            if not all(col in df.columns for col in expected_columns):
                # 列名の代替パターンをチェック
                column_mapping = {}
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'コンテンツ' in col or 'content' in col_lower:
                        column_mapping['コンテンツ名'] = col
                    elif '実績' in col or 'performance' in col_lower:
                        column_mapping['実績'] = col
                    elif '情報提供料' in col or 'information' in col_lower or 'fee' in col_lower:
                        column_mapping['情報提供料'] = col
                
                if len(column_mapping) < 3:
                    result.add_error(f"必要な列が見つかりません。検出された列: {list(df.columns)}")
                    return result
                
                # 列名をマッピング
                df = df.rename(columns=column_mapping)
            
            # 各行を処理してContentDetailを作成
            for _, row in df.iterrows():
                content_name = row.get('コンテンツ名', '')
                performance = pd.to_numeric(row.get('実績', 0), errors='coerce')
                info_fee = pd.to_numeric(row.get('情報提供料', 0), errors='coerce')
                sales_count = pd.to_numeric(row.get('売上件数', 1), errors='coerce') if '売上件数' in row else 1
                
                if pd.notna(content_name) and str(content_name).strip() and (performance > 0 or info_fee > 0):
                    detail = ContentDetail(
                        content_group=str(content_name).strip(),
                        performance=int(performance) if pd.notna(performance) else 0,
                        information_fee=int(info_fee) if pd.notna(info_fee) else 0,
                        sales_count=int(sales_count) if pd.notna(sales_count) else 1
                    )
                    result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'content_count': len(result.details),
                'file_type': '集計済み',
                'total_rows': len(df)
            }
            
            self.logger.info(f"集計済みLINE処理完了: {len(result.details)}コンテンツ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_docomo_file(self, file_path: Path) -> ProcessingResult:
        """docomo占いファイルを処理"""
        result = ProcessingResult(
            platform="docomo",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # bp40000746を含むファイルのみ処理
            if 'bp40000746' not in file_path.name:
                result.add_error("bp40000746を含むファイルではありません")
                self.logger.warning(f"docomo占いファイル処理スキップ: {file_path.name}")
                return result
            
            # ファイル形式に応じて読み込み
            if file_path.suffix.lower() == '.csv':
                # CSVファイルを読み込み（5行目以降を使用）
                df = self.csv_handler.read_csv_with_encoding_detection(
                    file_path,
                    skiprows=4,      # 先頭4行をスキップし5行目以降を読み込み対象
                    header=None      # 5行目をデータ行として扱う（ヘッダーなし）
                )
            elif file_path.suffix.lower() == '.pdf':
                # PDFファイルの場合はスキップ（CSVのみ処理）
                result.add_error("PDFファイルは現在サポートされていません")
                return result
            else:
                result.add_error(f"サポートされていないファイル形式: {file_path.suffix}")
                return result
            
            if df is None or df.empty:
                result.add_error("ファイルの読み込みに失敗またはデータが空です")
                return result
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # 列数チェック（R列=18列目、BI列=61列目、DK列=115列目が必要）
            if len(df.columns) < 115:
                result.add_error(f"列数が不足: 必要115列以上、実際{len(df.columns)}列")
                return result
            
            # R列（コンテンツ名）、BI列（実績）、DK列（情報提供料）を取得
            r_column = df.iloc[:, 17]  # R列（18番目、0ベースで17）
            bi_column = df.iloc[:, 60]  # BI列（61番目、0ベースで60）
            dk_column = df.iloc[:, 114]  # DK列（115番目、0ベースで114）
            
            # 数値に変換（BI列とDK列）
            bi_column = pd.to_numeric(bi_column, errors='coerce').fillna(0)
            dk_column = pd.to_numeric(dk_column, errors='coerce').fillna(0)
            
            # コンテンツ名でグループ化
            content_groups = {}
            for i, content_name in enumerate(r_column):
                if pd.notna(content_name) and str(content_name).strip():
                    content_name = str(content_name).strip()
                    if content_name not in content_groups:
                        content_groups[content_name] = {'bi_values': [], 'dk_values': []}
                    content_groups[content_name]['bi_values'].append(bi_column.iloc[i])
                    content_groups[content_name]['dk_values'].append(dk_column.iloc[i])
            
            # KEIKOソウルメイト占術の統合処理
            keiko_related_groups = {}
            other_groups = {}
            
            for content_name, values in content_groups.items():
                # KEIKOソウルメイト占術関連かチェック
                if ('ＫＥＩＫＯ' in content_name and 'ソウルメイト' in content_name) or \
                   ('KEIKO' in content_name and 'ソウルメイト' in content_name):
                    keiko_related_groups[content_name] = values
                else:
                    other_groups[content_name] = values
            
            # KEIKOソウルメイト占術関連の統合処理
            if keiko_related_groups:
                total_bi_values = []
                total_dk_values = []
                
                for content_name, values in keiko_related_groups.items():
                    total_bi_values.extend(values['bi_values'])
                    total_dk_values.extend(values['dk_values'])
                    self.logger.info(f"KEIKO統合対象: {content_name}")
                
                bi_sum = sum(total_bi_values)
                dk_sum = sum(total_dk_values)
                
                # BI列を1.1で除算したものが「実績」
                実績_sum = bi_sum / 1.1 if bi_sum > 0 else 0
                # DK列を1.1で除算したものが「情報提供料」
                情報提供料_sum = dk_sum / 1.1 if dk_sum > 0 else 0
                
                detail = ContentDetail(
                    content_group=self.DOCOMO_KEIKO_UNIFIED_NAME,
                    performance=round(実績_sum),
                    information_fee=round(情報提供料_sum),
                    sales_count=len(total_bi_values)  # 件数を追加
                )
                result.add_detail(detail)
                
                self.logger.info(f"KEIKO統合完了: {len(keiko_related_groups)}種類のコンテンツを統合")
            
            # その他のコンテンツの処理
            for content_name, values in other_groups.items():
                bi_sum = sum(values['bi_values'])
                dk_sum = sum(values['dk_values'])
                
                # BI列を1.1で除算したものが「実績」
                実績_sum = bi_sum / 1.1 if bi_sum > 0 else 0
                # DK列を1.1で除算したものが「情報提供料」
                情報提供料_sum = dk_sum / 1.1 if dk_sum > 0 else 0
                
                detail = ContentDetail(
                    content_group=content_name,
                    performance=round(実績_sum),
                    information_fee=round(情報提供料_sum),
                    sales_count=len(values['bi_values'])  # 件数を追加
                )
                result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'content_groups_count': len(content_groups),
                'total_rows': len(df)
            }
            
            self.logger.info(f"docomo処理完了: {len(content_groups)}コンテンツグループ")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_au_new_file(self, file_path: Path) -> ProcessingResult:
        """au占いファイルを処理（新仕様）"""
        result = ProcessingResult(
            platform="au",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # cp02お支払い明細書を含むファイル（PDFまたはCSV）のみ処理
            if 'cp02お支払い明細書' not in file_path.name or file_path.suffix.lower() not in ['.pdf', '.csv']:
                result.add_error("cp02お支払い明細書ファイル（PDFまたはCSV）ではありません")
                self.logger.warning(f"au占いファイル処理スキップ: {file_path.name}")
                return result
            
            # ファイル形式に応じて処理
            if file_path.suffix.lower() == '.csv':
                # auCSVファイル専用の読み込み処理（KEIKO占術専用）
                try:
                    # 特定セル（F9, N15）を取得するため直接CSVを解析
                    with open(file_path, 'r', encoding='shift_jis') as f:
                        lines = f.readlines()
                    
                    # F9セル（6列目、9行目）とN15セル（14列目、15行目）を直接取得
                    f9_value = None
                    n15_value = None
                    
                    # 9行目（インデックス8）のF列（インデックス5）を取得
                    if len(lines) > 8:
                        f9_line = lines[8].strip().split(',')
                        f9_fields = [field.strip('"') for field in f9_line]
                        if len(f9_fields) > 5:
                            try:
                                f9_value = float(f9_fields[5]) if f9_fields[5] else None
                            except ValueError:
                                pass
                    
                    # 15行目（インデックス14）のN列（インデックス13）を取得
                    if len(lines) > 14:
                        n15_line = lines[14].strip().split(',')
                        n15_fields = [field.strip('"') for field in n15_line]
                        if len(n15_fields) > 13:
                            try:
                                n15_value = float(n15_fields[13]) if n15_fields[13] else None
                            except ValueError:
                                pass
                    
                    if f9_value is None or n15_value is None:
                        result.add_error(f"必要なセル値が取得できません - F9: {f9_value}, N15: {n15_value}")
                        return result
                    
                    self.logger.info(f"auCSV セル値取得成功: F9={f9_value}, N15={n15_value}")
                    
                    # au占いの正しい計算ルール
                    # F9セルを1.1で除算した値が「実績」
                    実績 = f9_value / 1.1
                    # N15セルを1.1で除算した値が「情報提供料」
                    情報提供料 = n15_value / 1.1
                    
                    amounts_found = 2  # F9とN15の2つのセル値を取得
                    
                except Exception as e:
                    result.add_error(f"auCSV専用処理エラー: {str(e)}")
                    return result
                
            else:
                # PDFファイルの処理
                try:
                    import PyPDF2
                    import re
                    
                    text_content = ""
                    with open(file_path, 'rb') as pdf_file:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        for page in pdf_reader.pages:
                            try:
                                text_content += page.extract_text()
                            except Exception as e:
                                self.logger.warning(f"ページ読み込みエラー: {str(e)}")
                                continue
                    
                    # 改善された金額抽出パターン（auファイル用）
                    amount_patterns = [
                        r'合計金額[:：\s]*(\d{1,3}(?:,\d{3})*)',
                        r'金額[:：\s]*(\d{1,3}(?:,\d{3})*)', 
                        r'合計[:：\s]*(\d{1,3}(?:,\d{3})*)',
                        r'(\d{1,3}(?:,\d{3})*)[円\s]*$',  # 行末の金額
                        r'(\d{1,3}(?:,\d{3})*)'  # カンマ区切りの数字
                    ]
                    
                    amounts = []
                    for pattern in amount_patterns:
                        matches = re.findall(pattern, text_content, re.MULTILINE)
                        if matches:
                            amounts = matches
                            self.logger.info(f"au金額抽出成功 パターン: {pattern}, 結果: {matches[:3]}")
                            break
                    
                    if not amounts:
                        result.add_error("金額が見つかりません")
                        return result
                    
                    # 数値に変換（カンマを除去）
                    valid_amounts = []
                    for amount_str in amounts:
                        try:
                            amount_numeric = float(str(amount_str).replace(',', ''))
                            if amount_numeric > 100:  # 妥当な金額のみ（100円以上）
                                valid_amounts.append(amount_numeric)
                        except ValueError:
                            continue
                    
                    if not valid_amounts:
                        result.add_error("妥当な金額が見つかりません - 手動確認が必要です")
                        return result
                    
                    # 最大の金額を使用（通常は合計金額）
                    total_amount = max(valid_amounts)
                    
                except ImportError:
                    result.add_error("PyPDF2が必要です。pip install PyPDF2でインストールしてください")
                    return result
                except Exception as e:
                    result.add_error(f"PDF読み込みエラー: {str(e)}")
                    return result
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # CSVの場合はF9とN15から計算済み、PDFの場合はtotal_amountを使用
            if file_path.suffix.lower() == '.csv':
                # 既にF9とN15セルから計算済み
                pass
            else:
                # PDFの場合の処理
                if total_amount == 0:
                    result.add_error("有効な金額が見つかりません - 手動確認が必要です")
                    return result
                
                # PDFファイルの場合は従来通りの計算
                実績 = total_amount / 1.1
                情報提供料 = 実績 / 1.1
            
            detail = ContentDetail(
                content_group="ＫＥＩＫＯ☆ソウルメイト☆ソウルメイト占術",
                performance=round(実績),
                information_fee=round(情報提供料),
                sales_count=1  # auは1ファイル1件として扱う
            )
            result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            # CSVとPDF処理で適切な値を設定
            if file_path.suffix.lower() == '.csv':
                amounts_count = amounts_found
                result.metadata = {
                    'f9_value': f9_value,
                    'n15_value': n15_value,
                    'amounts_found': amounts_count,
                    'calculated_performance': round(実績),
                    'calculated_information_fee': round(情報提供料)
                }
            else:
                amounts_count = len(amounts) if 'amounts' in locals() else 0
                result.metadata = {
                    'total_amount': total_amount,
                    'amounts_found': amounts_count,
                    'calculated_performance': round(実績),
                    'calculated_information_fee': round(情報提供料)
                }
            
            self.logger.info(f"au処理完了: 実績={round(実績)}, 情報提供料={round(情報提供料)}")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def process_softbank_file(self, file_path: Path) -> ProcessingResult:
        """softbank占いファイルを処理（PDF）"""
        result = ProcessingResult(
            platform="softbank",
            file_name=file_path.name,
            success=False
        )
        
        start_time = datetime.now()
        
        try:
            # OID_PAY_9ATIを含むPDFファイルのみ処理
            if 'oid_pay_9ati' not in file_path.name.lower() or file_path.suffix.lower() != '.pdf':
                result.add_error("OID_PAY_9ATIを含むPDFファイルではありません")
                self.logger.warning(f"softbank占いファイル処理スキップ: {file_path.name}")
                return result
            
            # ファイル名から年月を抽出（OID_PAY_9ATI_yyyymm.PDF形式）
            year_month = self._extract_year_month_from_softbank_filename(file_path.name)
            if not year_month:
                result.add_error("ファイル名から年月を取得できません（OID_PAY_9ATI_yyyymm.PDF形式が必要）")
                return result
            
            # PDFファイルからテキストを抽出
            try:
                import PyPDF2
                import re
                
                text_content = ""
                with open(file_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    for page in pdf_reader.pages:
                        text_content += page.extract_text()
                
            except ImportError:
                result.add_error("PyPDF2が必要です。pip install PyPDF2でインストールしてください")
                return result
            except Exception as e:
                result.add_error(f"PDF読み込みエラー: {str(e)}")
                return result
            
            self.logger.log_file_operation("読み込み", file_path, True)
            
            # 改善された金額抽出パターン（softbankファイル用）
            # テスト結果から「8,176」のような金額を抽出できるパターンを使用
            amount_patterns = [
                r'(\d{1,3}(?:,\d{3})*)[円～\s]*',  # カンマ区切りの数字
                r'合計金額[:：\s]*(\d{1,3}(?:,\d{3})*)',
                r'お支払い金額[:：\s]*(\d{1,3}(?:,\d{3})*)',
                r'(\d+)'  # 数字のみ
            ]
            
            # 最低1つの妥当な金額を抽出
            found_amounts = []
            for pattern in amount_patterns:
                matches = re.findall(pattern, text_content, re.MULTILINE)
                for match in matches:
                    try:
                        amount_numeric = float(str(match).replace(',', ''))
                        if amount_numeric > 1000:  # 妥当な金額（1000円以上）
                            found_amounts.append(amount_numeric)
                    except ValueError:
                        continue
                if found_amounts:
                    self.logger.info(f"softbank金額抽出成功 パターン: {pattern}, 結果: {matches[:3]}")
                    break
            
            if not found_amounts:
                result.add_error("妥当な金額が見つかりません")
                return result
            
            # 最初の妥当な金額を使用（通常は最大の金額）
            main_amount = max(found_amounts)
            
            # softbankファイルの仕様に基づく計算
            # 金額をそのまま実績とし、実績の30%を情報提供料とする（仮の計算）
            total_sum = main_amount
            payment_sum = main_amount  # 同じ値を使用
            
            # 「合計金額」の合計に1.1を除算した値が「実績」
            実績 = total_sum / 1.1
            # 「お支払い金額」から1.1を除算したものが「情報提供料」
            情報提供料 = payment_sum / 1.1
            
            detail = ContentDetail(
                content_group="ＫＥＩＫＯソウルメイト占術",
                performance=round(実績),
                information_fee=round(情報提供料),
                sales_count=1  # softbankは1ファイル1件として扱う
            )
            result.add_detail(detail)
            
            # 合計を計算
            result.calculate_totals()
            result.success = True
            result.metadata = {
                'amounts_found': len(found_amounts),
                'main_amount': main_amount,
                'total_sum': total_sum,
                'payment_sum': payment_sum,
                'calculated_performance': round(実績),
                'calculated_information_fee': round(情報提供料),
                'year_month': year_month
            }
            
            self.logger.info(f"softbank処理完了: 年月={year_month}, 実績={round(実績)}, 情報提供料={round(情報提供料)}")
            
        except Exception as e:
            result.add_error(str(e))
            self.error_handler.handle_file_processing_error(e, file_path)
        
        finally:
            end_time = datetime.now()
            result.processing_time = (end_time - start_time).total_seconds()
        
        return result
    
    def _extract_year_month_from_softbank_filename(self, filename: str) -> str:
        """softbankファイル名から年月を抽出（OID_PAY_9ATI_yyyymm.PDF形式）"""
        try:
            # OID_PAY_9ATI_yyyymm.PDF形式から年月を抽出
            match = re.search(r'OID_PAY_9ATI_(\d{6})\.PDF', filename, re.IGNORECASE)
            if match:
                return match.group(1)
            
            # 見つからない場合は空文字を返す
            return ""
            
        except Exception as e:
            self.logger.warning(f"softbank年月抽出エラー: {filename} - {str(e)}")
            return ""
    
    def _extract_year_month_from_path(self, file_path: Path) -> str:
        """ファイルのパスから年月を抽出（YYYYMM形式）"""
        try:
            # パスの親フォルダ（年月フォルダ）名から6桁の年月を取得
            for parent in file_path.parents:
                folder_name = parent.name
                if re.match(r'\d{6}', folder_name):
                    return folder_name
            
            # フォルダ名から取得できない場合、ファイル名から推測
            filename = file_path.name
            # ファイル名に含まれる年月パターンを検索（例：202302）
            match = re.search(r'(\d{4})(\d{2})', filename)
            if match:
                year = match.group(1)
                month = match.group(2)
                return f"{year}{month}"
            
            # 見つからない場合は空文字を返す
            return ""
            
        except Exception as e:
            self.logger.warning(f"年月抽出エラー: {file_path} - {str(e)}")
            return ""
    
    def process_all_files(self):
        """すべてのファイルを一括処理"""
        self.logger.info("全ファイル処理を開始")
        
        # 年月フォルダ内のファイルを検索
        files_by_platform = self.find_files_in_yearmonth_folders()
        
        total_files = sum(len(files) for files in files_by_platform.values())
        self.logger.info(f"処理対象ファイル数: {total_files}")
        
        # プラットフォーム別にファイルを処理
        for platform, files in files_by_platform.items():
            for file_path in files:
                self.logger.info(f"処理中: {platform} - {file_path.name}")
                
                if platform == 'ameba':
                    result = self.process_ameba_file(file_path)
                elif platform == 'rakuten':
                    result = self.process_rakuten_file(file_path)
                elif platform == 'mediba':
                    result = self.process_au_file(file_path)
                elif platform == 'excite':
                    result = self.process_excite_file(file_path)
                elif platform == 'line':
                    result = self.process_line_file(file_path)
                elif platform == 'docomo':
                    result = self.process_docomo_file(file_path)
                elif platform == 'au':
                    result = self.process_au_new_file(file_path)
                elif platform == 'softbank':
                    result = self.process_softbank_file(file_path)
                else:
                    self.logger.warning(f"未対応プラットフォーム: {platform}")
                    continue
                
                if result.success:
                    # softbankの場合はファイル名から取得した年月を使用、それ以外はパスから取得
                    if platform == 'softbank' and result.metadata and 'year_month' in result.metadata:
                        year_month = result.metadata['year_month']
                    else:
                        year_month = self._extract_year_month_from_path(file_path)
                    
                    self.results.append({
                        'platform': result.platform,
                        'file_name': result.file_name,
                        'content_details': result.details,
                        '情報提供料': result.total_information_fee,
                        '実績合計': result.total_performance,
                        '年月': year_month,
                        '処理日時': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    self.logger.info(f"処理成功: {file_path.name}")
                else:
                    self.logger.error(f"処理失敗: {file_path.name} - {', '.join(result.errors)}")
        
        self.logger.info(f"全ファイル処理完了: {len(self.results)}件成功")
        
        # docomoのKEIKOソウルメイト占術の結果を合算
        self._consolidate_docomo_keiko_results()
    
    def _consolidate_docomo_keiko_results(self):
        """docomoのKEIKOソウルメイト占術の結果を年月別に合算（同一プラットフォーム内のみ）"""
        if not self.results:
            return
        
        # docomo内のKEIKOソウルメイト占術データの統計情報を出力
        for result in self.results:
            if result['platform'] != 'docomo':
                continue
                
            year_month = result.get('年月', '')
            
            # docomoのKEIKOソウルメイト占術の詳細データを検索
            if result['content_details']:
                keiko_details = [detail for detail in result['content_details'] 
                               if detail.content_group == self.DOCOMO_KEIKO_UNIFIED_NAME]
                
                if keiko_details:
                    total_performance = sum(detail.performance for detail in keiko_details)
                    total_information_fee = sum(detail.information_fee for detail in keiko_details)
                    
                    self.logger.info(f"docomo KEIKOソウルメイト占術統合結果 {year_month}: "
                                   f"実績={total_performance:,}円, "
                                   f"情報提供料={total_information_fee:,}円, "
                                   f"ファイル={result['file_name']}")
    
    def export_to_csv(self, output_path: str):
        """結果をCSVファイルに出力（重複除去機能付き）"""
        if not self.results:
            self.logger.warning("出力するデータがありません")
            return
        
        try:
            # 重複除去を行う
            deduplicated_data = self._deduplicate_data()
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # ヘッダー
                writer.writerow(['プラットフォーム', 'ファイル名', 'コンテンツ', '実績', '情報提供料', '売上件数', '年月', '処理日時'])
                
                # 重複除去されたデータを出力
                for row_data in deduplicated_data:
                    writer.writerow(row_data)
            
            self.logger.info(f"CSV出力完了: {output_path}")
            
        except Exception as e:
            self.logger.error(f"CSV出力エラー: {str(e)}")
            raise
    
    def _deduplicate_data(self) -> list:
        """データの重複除去を行う"""
        seen_keys = set()
        deduplicated_data = []
        duplicate_count = 0
        
        for result in self.results:
            platform = result['platform']
            file_name = result['file_name']
            year_month = result.get('年月', '')
            processing_time = result.get('処理日時', '')
            
            if result['content_details']:
                for detail in result['content_details']:
                    # 重複チェック用のキーを作成（プラットフォーム、コンテンツ、年月の組み合わせ）
                    key = (platform, detail.content_group, year_month)
                    
                    if key not in seen_keys:
                        seen_keys.add(key)
                        deduplicated_data.append([
                            platform,
                            file_name,
                            detail.content_group,
                            detail.performance,
                            detail.information_fee,
                            detail.sales_count,
                            year_month,
                            processing_time
                        ])
                    else:
                        duplicate_count += 1
                        self.logger.warning(f"重複データを除去: {platform} - {detail.content_group} - {year_month}")
            else:
                # 詳細がない場合は合計値を出力
                key = (platform, '合計', year_month)
                if key not in seen_keys:
                    seen_keys.add(key)
                    deduplicated_data.append([
                        platform,
                        file_name,
                        '合計',
                        result['実績合計'],
                        result['情報提供料'],
                        0,  # 合計行では件数を0とする
                        year_month,
                        processing_time
                    ])
                else:
                    duplicate_count += 1
                    self.logger.warning(f"重複データを除去: {platform} - 合計 - {year_month}")
        
        if duplicate_count > 0:
            self.logger.info(f"重複データ除去完了: {duplicate_count}件の重複を除去")
        
        return deduplicated_data