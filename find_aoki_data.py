#!/usr/bin/env python3
"""
aokiの実際のデータを探すスクリプト
"""

import pandas as pd
import openpyxl
from pathlib import Path

def find_aoki_data():
    """202505_aoki.xlsxで実際に計算されるべき値を探す"""
    excel_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    if not Path(excel_path).exists():
        print(f"ファイルが見つかりません: {excel_path}")
        return
    
    # Excelファイルを開く
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    print("=== 202505_aoki.xlsx の全データ確認 ===")
    
    # 全てのセルで非ゼロ値を探す
    print("\n--- 非ゼロ数値を含むセル ---")
    found_numbers = []
    target_values = [58, 81, 163]
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            
            # 数値の場合
            if isinstance(cell_value, (int, float)) and cell_value != 0:
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                found_numbers.append((cell_ref, cell_value))
                
                if cell_value in target_values:
                    print(f"  🎯 {cell_ref}: {cell_value} (目標値)")
                else:
                    print(f"     {cell_ref}: {cell_value}")
    
    print(f"\n見つかった数値セル総数: {len(found_numbers)}")
    
    # 目標値の合計を計算
    found_targets = [(ref, val) for ref, val in found_numbers if val in target_values]
    if found_targets:
        print(f"\n目標値 58, 81, 163 が見つかりました:")
        for ref, val in found_targets:
            print(f"  {ref}: {val}")
        print(f"合計: {sum([val for _, val in found_targets])}")
    
    # 式の値を手動計算してみる
    print(f"\n--- M19セルの数式 =SUM(AE23:AJ58) の詳細確認 ---")
    
    # 数式版でExcelを開き直し
    wb_formula = openpyxl.load_workbook(excel_path, data_only=False)
    ws_formula = wb_formula.active
    
    total_manual_sum = 0
    
    for row in range(23, 59):  # 23から58まで
        for col_letter in ['AE', 'AF', 'AG', 'AH', 'AI', 'AJ']:
            cell_ref = f"{col_letter}{row}"
            
            # 数式版の値
            formula_cell = ws_formula[cell_ref]
            # データ版の値
            data_cell = ws[cell_ref]
            
            if data_cell.value and data_cell.value != 0:
                print(f"  {cell_ref}: 値={data_cell.value}, 数式={formula_cell.value}")
                if isinstance(data_cell.value, (int, float)):
                    total_manual_sum += data_cell.value
    
    print(f"\n手動計算したSUM(AE23:AJ58)の合計: {total_manual_sum}")
    
    # M19セルの実際の値
    m19_value = ws['M19'].value
    print(f"M19セルの実際の値: {m19_value}")
    
    # 直接計算してみる（Excelを使わずに）
    print(f"\n--- 直接計算による検証 ---")
    # まず、何行目にデータが入っているかを確認
    print("データが入っている行を探します:")
    for row in range(20, 70):
        row_has_data = False
        row_data = {}
        
        for col_letter in ['A', 'D', 'G', 'M', 'S', 'Y', 'AC', 'AE']:
            cell_ref = f"{col_letter}{row}"
            cell_value = ws[cell_ref].value
            
            # 数式版の値も確認（AE列の場合）
            if col_letter == 'AE':
                formula_cell = ws_formula[cell_ref]
                if formula_cell.value:
                    row_data[f'{col_letter}_formula'] = formula_cell.value
            
            if cell_value is not None and cell_value != "" and cell_value != 0:
                row_data[col_letter] = cell_value
                row_has_data = True
        
        if row_has_data:
            print(f"  行{row}: {row_data}")

if __name__ == '__main__':
    find_aoki_data()