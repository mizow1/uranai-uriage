#!/usr/bin/env python3
"""
Excelファイルを開いて強制的に再計算させるスクリプト
"""

import openpyxl
from pathlib import Path
import win32com.client
import sys

def force_excel_calculation():
    """Excelファイルを開いて強制的に再計算"""
    excel_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    if not Path(excel_path).exists():
        print(f"ファイルが見つかりません: {excel_path}")
        return
    
    try:
        # COMアプリケーションでExcelを開く
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False  # バックグラウンドで実行
        
        # ファイルを開く
        workbook = excel_app.Workbooks.Open(excel_path)
        worksheet = workbook.ActiveSheet
        
        # 計算を強制実行
        excel_app.Calculate()
        
        # M19セルの値を取得
        m19_value = worksheet.Range("M19").Value
        print(f"M19セルの計算結果: {m19_value}")
        
        # AE列の値も確認
        print("AE列の計算結果:")
        for row in [23, 25, 26]:
            ae_value = worksheet.Range(f"AE{row}").Value
            y_value = worksheet.Range(f"Y{row}").Value
            ac_value = worksheet.Range(f"AC{row}").Value
            print(f"  AE{row}: {ae_value} (Y{row}={y_value}, AC{row}={ac_value})")
        
        # ファイルを保存
        workbook.Save()
        
        # 閉じる
        workbook.Close()
        excel_app.Quit()
        
        print(f"計算完了。M19の値: {m19_value}")
        
    except Exception as e:
        print(f"エラー: {e}")
        # Excel が残っている場合はクリーンアップ
        try:
            excel_app.Quit()
        except:
            pass

def check_with_openpyxl():
    """openpyxlで再度確認"""
    excel_path = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ\202505\202505_aoki.xlsx"
    
    print("\n=== openpyxlでの確認 ===")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    m19_value = ws['M19'].value
    print(f"M19セルの値: {m19_value}")
    
    total = 0
    for row in [23, 25, 26]:
        ae_value = ws[f'AE{row}'].value
        print(f"AE{row}: {ae_value}")
        if ae_value and isinstance(ae_value, (int, float)):
            total += ae_value
    
    print(f"AE列合計（手動）: {total}")

if __name__ == '__main__':
    try:
        force_excel_calculation()
    except ImportError:
        print("win32comが利用できません。openpyxlのみで確認します。")
    
    check_with_openpyxl()