#\!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AC列の数式を確認
"""

import openpyxl
import glob
import os

def check_ac_formulas(file_path):
    """AC列の数式を確認"""
    print(f"=== {os.path.basename(file_path)} のAC列分析 ===")
    
    try:
        workbook_formula = openpyxl.load_workbook(file_path, data_only=False)
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)
        
        worksheet_formula = workbook_formula.active
        worksheet_data = workbook_data.active
        
        print(f"AC列の数式と値（23-58行目）:")
        
        for row in range(23, 59):  # 23-58行目
            ac_formula = worksheet_formula.cell(row=row, column=29)  # AC列 = 29列目
            ac_data = worksheet_data.cell(row=row, column=29)
            
            if ac_formula.value is not None or ac_data.value is not None:
                print(f"  AC{row}:")
                print(f"    数式: {ac_formula.value}")
                print(f"    計算値: {ac_data.value}")
                print(f"    型: {type(ac_data.value)}")
                
                if isinstance(ac_formula.value, str) and ac_formula.value.startswith("="):
                    print(f"    -> これは数式です")
                print()
        
        workbook_formula.close()
        workbook_data.close()
        
    except Exception as e:
        print(f"エラー: {e}")

def main():
    # テスト用に1つのファイルを分析
    royalty_dir = r"C:\Users\OW\Dropbox\disk2とローカルの同期\占い\占い売上\履歴\ロイヤリティ"
    
    # 202508フォルダから1つのファイルを選択
    test_folder = os.path.join(royalty_dir, "202508")
    if os.path.exists(test_folder):
        xlsx_files = glob.glob(os.path.join(test_folder, "*.xlsx"))
        if xlsx_files:
            # 最初のファイルをテスト
            check_ac_formulas(xlsx_files[0])
        else:
            print("No xlsx files found")
    else:
        print("Test folder not found")

if __name__ == "__main__":
    main()
