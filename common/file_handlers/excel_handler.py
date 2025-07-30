"""
統一Excelハンドラー
"""
import pandas as pd
import openpyxl
from pathlib import Path
from typing import List, Optional, Dict, Any
import io
from ..error_handling.exceptions import FileProcessingError


class ExcelHandler:
    """Excelファイルの統一処理クラス"""
    
    DEFAULT_PASSWORDS = ['', 'password', '123456', '000000', 'admin', 'user']
    DEFAULT_ENGINES = ['openpyxl', 'xlrd']
    
    def __init__(self, logger=None, error_handler=None):
        self.logger = logger
        self.error_handler = error_handler
    
    def read_excel_with_password_handling(self, file_path: Path, passwords: Optional[List[str]] = None, **kwargs) -> pd.DataFrame:
        """パスワード保護処理付きでExcelファイルを読み込み"""
        if passwords is None:
            passwords = self.DEFAULT_PASSWORDS
        
        # まず通常の読み込みを試行
        try:
            df = pd.read_excel(file_path, **kwargs)
            if self.logger:
                self.logger.info(f"Excel読み込み成功（パスワード保護なし）: {file_path.name}")
            return df
        except Exception as e:
            if "password" in str(e).lower() or "protected" in str(e).lower() or "zip file" in str(e).lower():
                return self._try_password_protected(file_path, passwords, **kwargs)
            else:
                raise FileProcessingError(f"Excel読み込みエラー: {file_path.name} - {str(e)}")
    
    def _try_password_protected(self, file_path: Path, passwords: List[str], **kwargs) -> pd.DataFrame:
        """パスワード保護されたExcelファイルの解除を試行"""
        try:
            import msoffcrypto
        except ImportError:
            raise FileProcessingError(f"パスワード保護されたExcelファイルの処理にはmsoffcrypto-toolが必要です: {file_path.name}")
        
        for password in passwords:
            try:
                with open(file_path, 'rb') as file:
                    office_file = msoffcrypto.OfficeFile(file)
                    office_file.load_key(password=password if password else None)
                    
                    decrypted = io.BytesIO()
                    office_file.save(decrypted)
                    decrypted.seek(0)
                    
                    df = pd.read_excel(decrypted, **kwargs)
                    if self.logger:
                        self.logger.info(f"Excel読み込み成功（パスワード: '{password}'）: {file_path.name}")
                    return df
                    
            except Exception as e:
                if self.logger:
                    self.logger.debug(f"パスワード試行失敗: {file_path.name} ('{password}') - {str(e)}")
                continue
        
        raise FileProcessingError(f"すべてのパスワードでExcel解除に失敗: {file_path.name}")
    
    def try_multiple_engines(self, file_path: Path, engines: Optional[List[str]] = None, **kwargs) -> pd.DataFrame:
        """複数のExcelエンジンを順次試行"""
        if engines is None:
            engines = self.DEFAULT_ENGINES
        
        last_error = None
        
        for engine in engines:
            try:
                kwargs_with_engine = kwargs.copy()
                kwargs_with_engine['engine'] = engine
                
                df = pd.read_excel(file_path, **kwargs_with_engine)
                if self.logger:
                    self.logger.info(f"Excel読み込み成功: {file_path.name} (engine: {engine})")
                return df
                
            except Exception as e:
                last_error = e
                if self.logger:
                    self.logger.debug(f"Excel読み込み失敗: {file_path.name} (engine: {engine}) - {str(e)}")
                continue
        
        raise FileProcessingError(f"すべてのエンジンでExcel読み込みに失敗: {file_path.name} - 最後のエラー: {str(last_error)}")
    
    def handle_encrypted_files(self, file_path: Path, passwords: Optional[List[str]] = None, **kwargs) -> pd.DataFrame:
        """暗号化されたExcelファイルを処理"""
        return self.read_excel_with_password_handling(file_path, passwords, **kwargs)
    
    def read_excel_safe(self, file_path: Path, **kwargs) -> Optional[pd.DataFrame]:
        """安全なExcel読み込み（エラー時はNoneを返す）"""
        try:
            return self.read_excel_with_password_handling(file_path, **kwargs)
        except Exception as e:
            if self.error_handler:
                self.error_handler.handle_file_processing_error(e, file_path)
            elif self.logger:
                self.logger.error(f"Excel読み込みエラー: {file_path.name} - {str(e)}")
            return None
    
    def get_sheet_names(self, file_path: Path, passwords: Optional[List[str]] = None) -> List[str]:
        """Excelファイルのシート名一覧を取得"""
        try:
            # パスワード保護の場合の処理
            if passwords is None:
                passwords = self.DEFAULT_PASSWORDS
            
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                return wb.sheetnames
            except Exception as e:
                if "password" in str(e).lower() or "protected" in str(e).lower():
                    return self._get_sheet_names_encrypted(file_path, passwords)
                else:
                    raise
                    
        except Exception as e:
            if self.logger:
                self.logger.error(f"シート名取得エラー: {file_path.name} - {str(e)}")
            return []
    
    def _get_sheet_names_encrypted(self, file_path: Path, passwords: List[str]) -> List[str]:
        """暗号化されたExcelファイルからシート名を取得"""
        try:
            import msoffcrypto
        except ImportError:
            if self.logger:
                self.logger.error("msoffcrypto-toolが必要です")
            return []
        
        for password in passwords:
            try:
                with open(file_path, 'rb') as file:
                    office_file = msoffcrypto.OfficeFile(file)
                    office_file.load_key(password=password if password else None)
                    
                    decrypted = io.BytesIO()
                    office_file.save(decrypted)
                    decrypted.seek(0)
                    
                    wb = openpyxl.load_workbook(decrypted, data_only=True)
                    return wb.sheetnames
                    
            except Exception:
                continue
        
        return []
    
    def get_file_info(self, file_path: Path) -> Dict[str, Any]:
        """Excelファイルの基本情報を取得"""
        try:
            sheet_names = self.get_sheet_names(file_path)
            
            return {
                'file_name': file_path.name,
                'sheet_names': sheet_names,
                'sheet_count': len(sheet_names),
                'file_size': file_path.stat().st_size
            }
        except Exception as e:
            if self.logger:
                self.logger.error(f"Excelファイル情報取得エラー: {file_path.name} - {str(e)}")
            return {}